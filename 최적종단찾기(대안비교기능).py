import random
import copy
import math
import csv
import matplotlib.pyplot as plt
from tkinter import filedialog, Tk, StringVar, ttk
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk  # 추가
import numpy as np
import openpyxl
import tkinter as tk
from matplotlib.figure import Figure

# 전역변수
chain = 25
isstartend = True  # True 이면 존재 False 면 미존재
destination_score = 100  # 목표점수
similarity_threshold = 5  # 예측 정확도

start_break_point = 1000  # 시작 vip고정
end_break_point = 11550  # 종점 vip고정

plt.rcParams['font.family'] ='Malgun Gothic'
plt.rcParams['axes.unicode_minus'] =False

# 생성할 대안 갯수
alternative_count = 'inf'

if isstartend:
    detail_design = [(0, 51), (12350, 46)]
else:
    detail_design = [(0, 0)]

def gui_thread():
    root = tk.Tk()
    label = tk.Label(root, text="GUI 메인 루프를 실행합니다.")
    label.pack()
    root.mainloop()
    
# csv 파일에서 지반고 정보와 측점 정보를 불러와서 gl 배열에 저장
def load_gl():
    gl = []
    file_path = filedialog.askopenfilename(title="파일 선택", filetypes=[("텍스트 파일", "*.txt")])

    if file_path:
        try:
            with open(file_path, 'r', encoding='UTF-8') as file:
                reader = csv.reader(file)
                next(reader)  # 첫 번째 행 건너뛰기
                for row in reader:
                    station = int(row[0])  # 측점 정보
                    ground_elevation = float(row[1])  # 지반고 정보
                    
                    gl.append([station, ground_elevation])
                    
        except FileNotFoundError:
            print("파일을 찾을 수 없습니다.")
        except Exception as e:
            print("파일을 읽는 중 오류가 발생했습니다:", e)
    else:
        print("파일 선택이 취소되었습니다.")

    return gl

def detect_terrain(gl):
    # 경사도를 저장할 리스트
    slope_type = []

    # 고도 데이터를 반복하여 경사도를 계산
    for i in range(1, len(gl) - 1):
        prev_elevation = gl[i - 1][1]
        current_elevation = gl[i][1]
        next_elevation = gl[i + 1][1]

        # 현재 고도와 이전, 다음 고도의 차이를 계산
        diff_prev = current_elevation - prev_elevation
        diff_next = next_elevation - current_elevation

        # 거리 변화가 0이 아닌 경우에만 경사도를 계산하여 추가
        if diff_prev != 0 and diff_next != 0:
            slope_prev = (diff_prev / chain) * 100  # 이전 측점에서 현재 측점까지의 경사도
            slope_next = (diff_next / chain) * 100  # 현재 측점에서 다음 측점까지의 경사도
            average_slope = (slope_prev + slope_next) / 2  # 두 경사도의 평균을 사용하여 현재 측점의 경사도 계산

            # 평균 경사도에 따라 경사도 유형을 할당
            if average_slope < 3:
                slope_type.append('평지')
            elif 3 <= average_slope < 10:
                slope_type.append('완만한 경사')
            elif 10 <= average_slope < 20:
                slope_type.append('중간 경사')
            else:
                slope_type.append('급한 경사')
        else:
            slope_type.append('N/A')  # 경사도를 계산할 수 없는 경우

    return slope_type

def generate_random_profile(num_points, min_distance, gl):
    # 초기 설정
    start_station, start_elevation = gl[0]
    end_station, end_elevation = gl[-1]

    detail_start_station, detail_start_elevation = detail_design[0]
    detail_end_station, detail_end_elevation = detail_design[-1]
    
    # 시작점 설정
    if not isstartend:  # False
        points = [[start_station, start_elevation + 10]]  # 초기 계획고 설정
    else:
        points = [[detail_start_station, detail_start_elevation]]

    # 현재 위치 설정
    current_station = start_station
    current_elevation = start_elevation + 10

    for i in range(num_points - 1):
        # 다음 지점까지의 거리 설정
        distance_to_next = chain * math.ceil(random.uniform(min_distance, min_distance * 2) / chain)

        if current_station + distance_to_next >= end_station:
            break

        # 다음 지점 설정
        next_station = current_station + distance_to_next
        
        # 다음 지점의 지반고 찾기
        next_elevation = None

        if isstartend:
            for station, elevation in gl:
                if station == next_station:
                    next_elevation = elevation + 10
                    break
            if i == 0:
                next_elevation = detail_start_elevation
        else:
            for station, elevation in gl:
                if station == next_station:
                    next_elevation = elevation + 10
                    break

        # 다음 지점의 지반고가 없는 경우 예외 처리
        if next_elevation is None:
            raise ValueError(f"No ground elevation data found for station {next_station}")

        # 현재 위치 업데이트
        current_station = next_station
        current_elevation = next_elevation

        # 지점 추가
        points.append([current_station, current_elevation])

    # 마지막 지점 설정
    if isstartend:
        points.append([end_station, detail_end_elevation])
        points[-2][0] = end_break_point
        points[-2][1] = points[-1][1]
    else:
        points.append([end_station, end_elevation + 10])
    return points

def check_and_adjust_elevation(profile):
    adjusted_profile = []
    for i, (station, elevation) in enumerate(profile):
        rand_el = random.uniform(0, 20)
        if isstartend:
            if i == 1:
                station = start_break_point
            elif i > 1 and i < len(profile) - 3:
                prev_station, prev_elevation = adjusted_profile[-1]
                if abs(elevation - prev_elevation) > 20:
                    elevation = prev_elevation + (rand_el if elevation > prev_elevation else -rand_el)

        else:
            if i > 0:
                prev_station, prev_elevation = adjusted_profile[-1]
                if abs(elevation - prev_elevation) > 20:
                    elevation = prev_elevation + (rand_el if elevation > prev_elevation else -rand_el)
        adjusted_profile.append([station, elevation])

    return adjusted_profile

# 최급기울기 찾기
def find_steepest_gradient(g):
    max_gradient = 0
    steepest_gradient_station = None
    
    for i in range(len(g) - 1):
        station1, elevation1 = g[i]
        station2, elevation2 = g[i + 1]
        
        # 두 점 사이의 거리
        distance = station2 - station1
        
        if distance == 0:
            continue
        
        # 기울기 계산
        gradient = abs((elevation2 - elevation1) / distance) * 1000
        
        # 최대 기울기 갱신
        if gradient > max_gradient:
            max_gradient = gradient
            steepest_gradient_station = station1 if elevation1 < elevation2 else station2
            
    return max_gradient, len(g) - 2

# 표고점을 20간격으로 계획고 생성
def generate_station_elv(points):
    station_elv = []

    for i in range(len(points) - 1):
        start_station = points[i][0]
        start_elevation = points[i][1]
        end_station = points[i + 1][0]
        end_elevation = points[i + 1][1]

        
        height = end_elevation - start_elevation
        distance = end_station - start_station

    
        if distance == 0:
            slope = 0
        else:
            slope = height / distance * 1000

        num_steps = int(distance / chain)  # 20의 배수인 측점 개수 계산

        for j in range(num_steps + 1):
            current_station = start_station + j * chain
            current_elevation = start_elevation + slope / 1000 * (current_station - start_station)
            station_elv.append([current_station, current_elevation])

    return station_elv

# 절토성토고 계산
def calculate_cutfill(station_elv, gl):
    new_cutfill = []
    
    for gl_station, el in gl:  # 지반고 추출
        for elv_station, dl in station_elv:  # 계획고 추출
            if gl_station == elv_station:
                height_d = el - dl
                new_cutfill.append([gl_station, height_d])

    return new_cutfill

# 구조물 판단
def structure_judgement(new_cutfill):
    new_structure = []

    for station, cutfill_d in new_cutfill:
        if cutfill_d >= 15:
            structure_type = '교량'
        elif cutfill_d <= -15:
            structure_type = '터널'
        elif cutfill_d > 0:
            structure_type = '절토'
        elif cutfill_d < 0:
            structure_type = '성토'
        else:
            structure_type = 'N/A'

        new_structure.append([station, structure_type])

    return new_structure

# 절성토량을 구하고 점수 매기기
def calculate_cutfill_volume_and_score(new_cutfill):
    total_cut_volume = 0
    total_fill_volume = 0

    for station, cutfill_d in new_cutfill:
        if cutfill_d > 0:
            total_cut_volume += cutfill_d
        elif cutfill_d < 0:
            total_fill_volume += abs(cutfill_d)
            
    average_score = (total_cut_volume + total_fill_volume) / 2
    
    return average_score, total_cut_volume, total_fill_volume

# 단순해 체크
def is_dominated(profile1, profile2):
    station_elv_1 = generate_station_elv(profile1)
    station_elv_2 = generate_station_elv(profile2)

    cutfill_1 = calculate_cutfill(station_elv_1, gl)
    cutfill_2 = calculate_cutfill(station_elv_2, gl)

    score1, total_cut_volume1, total_fill_volume1 = calculate_cutfill_volume_and_score(cutfill_1)
    score2, total_cut_volume2, total_fill_volume2 = calculate_cutfill_volume_and_score(cutfill_2)

    return score1 >= score2

#절토성토고 계산
def calculate_cutfill(station_elv, gl):
    new_cutfill = []
    
    for gl_station, el in gl:  # 지반고 추출
        for elv_station, dl in station_elv:  # 계획고 추출
            if gl_station == elv_station:
                height_d = el - dl
                new_cutfill.append([gl_station, height_d])

    return new_cutfill

#구조물판별
def cal_profile_structure(new_cutfill):
    new_structure = []

    for station, cutfill_d in new_cutfill:
        if cutfill_d  >= 12:
             new_structure.append([station , "터널"])
            
        elif cutfill_d  <= -12:
            new_structure.append([station , "교량"])
        else:
             new_structure.append([station , "토공"])

    return new_structure

# 공사비용 계산 함수
def cal_profile_money(new_structure):
    length_bridge = 0
    length_tunnel = 0
    length_excavation = 0

    for station, structure_type in new_structure:
        if "교량" in structure_type:
            length_bridge += chain
        if "터널" in structure_type:
            length_tunnel += chain
        if "토공" in structure_type:
            length_excavation += chain

    cost_bridge = (length_bridge / 1000) * 24574
    cost_tunnel = (length_tunnel / 1000) * 15784
    cost_excavation = (length_excavation / 1000) * 8620

    new_total_cost = cost_bridge + cost_tunnel + cost_excavation

    return new_total_cost

#구조물 측점
def find_structure_ranges(structure_list):
    ranges = {}
    bridge_count = 1
    tunnel_count = 1

    current_structure = None
    start_station = None

    for i, (station, structure) in enumerate(structure_list):
        if structure == '교량':
            if current_structure != '교량':
                if current_structure is not None:
                    if current_structure == '터널':
                        key = f'T{tunnel_count}'
                        ranges[key] = (start_station, structure_list[i - 1][0])
                        tunnel_count += 1
                start_station = station
                current_structure = '교량'
        elif structure == '터널':
            if current_structure != '터널':
                if current_structure is not None:
                    if current_structure == '교량':
                        key = f'B{bridge_count}'
                        ranges[key] = (start_station, structure_list[i - 1][0])
                        bridge_count += 1
                start_station = station
                current_structure = '터널'
        else:
            if current_structure == '교량':
                key = f'B{bridge_count}'
                ranges[key] = (start_station, structure_list[i - 1][0])
                bridge_count += 1
            elif current_structure == '터널':
                key = f'T{tunnel_count}'
                ranges[key] = (start_station, structure_list[i - 1][0])
                tunnel_count += 1
            current_structure = None
            start_station = None

    if current_structure == '교량':
        key = f'B{bridge_count}'
        ranges[key] = (start_station, structure_list[-1][0])
    elif current_structure == '터널':
        key = f'T{tunnel_count}'
        ranges[key] = (start_station, structure_list[-1][0])

    return ranges

# 구조물 갯수
def count_structures(structure_ranges):
    count_bridge = 0
    count_tunnel = 0

    for key in structure_ranges.keys():
        if key.startswith('B'):
            count_bridge += 1
        elif key.startswith('T'):
            count_tunnel += 1

    return count_bridge, count_tunnel

# Plotting and combo box integration
def plot_profile(profile, gl, label, color):
    station_elv = generate_station_elv(profile)
    stations = [station for station, elevation in profile]
    elevations = [elevation for station, elevation in profile]
    
    ax.plot(stations, elevations, label=label, color=color)
    ax.scatter([station for station, elevation in profile],
                [elevation for station, elevation in profile],
                color=color, s=10)
    #기울기 표시
    # 각 계획선의 중심에 기울기 표시
    i= 0
    for i in range(len(stations) - 1):
        x1, y1 = stations[i], elevations[i]
        x2, y2 = stations[i + 1], elevations[i + 1]
        
        
        gradient = (y2 - y1) / (x2 - x1) * 1000 # 기울기 계산
        if gradient !=0:
            gradient_text = f'{gradient:.2f}'.rstrip('0').rstrip('.')
        else:
            gradient_text = 'L'  # 분모가 0인 경우 'L'로 설정
        
        midpoint = (x1 + x2) / 2, (y1 + y2) / 2   + 10# 두 점 사이의 중심

        ax.text(midpoint[0], midpoint[1], gradient_text, fontsize=10, color=color, ha='center', va='center')

    cutfill = calculate_cutfill(station_elv, gl)
    # Calculate and print construction cost
    struct = cal_profile_structure(cutfill)
    # Determine and print structures
    struct_station = find_structure_ranges(struct)
    
    #구조물 시종점 표시
    # 교량 시종점에 수직선 추가
    for bridge_key, (start_station, end_station) in struct_station.items():
        for station, elevation in station_elv:
            if bridge_key.startswith('B') and start_station == station:
                ax.plot([start_station, start_station], [elevation, elevation + 100], color=color, linestyle='-')
                ax.text(start_station, elevation + 100, bridge_key, color=color)
            elif bridge_key.startswith('B') and end_station == station:
                ax.plot([end_station, end_station], [elevation, elevation + 100], color=color, linestyle='-')

            elif bridge_key.startswith('T') and start_station == station:
                ax.plot([start_station, start_station], [elevation, elevation + 100], color=color, linestyle='-')
                ax.text(start_station, elevation + 100, bridge_key, color=color)
            elif bridge_key.startswith('T') and end_station == station:
                ax.plot([end_station, end_station], [elevation, elevation + 100], color=color, linestyle='-')
            else:
                continue
            
    alternative = profile_spec(profile)
    optimal_alternative = profile_spec(optimal_profile)
    create_excel(optimal_alternative,'최적대안')
    create_excel(alternative,'대안')
    
def plot_ground_profile(gl):
    stations = [station for station, elevation in gl]
    elevations = [elevation for station, elevation in gl]
    
    ax.plot(stations, elevations, label='지반고', color='black')

# 최적 대안과 다른 대안들 표시
def plot_alternatives(optimal_profile, alternative_profiles, selected_profile=None):
    ax.clear()  # Clear the current plot
    plot_ground_profile(gl)
    plot_profile(optimal_profile, gl, '최적 대안', 'red')
    
    if selected_profile:
        plot_profile(selected_profile, gl, '선택된 대안', 'blue')
    
    ax.set_title('종단최적화 프로그램')
    ax.set_xlabel('X-axis')
    ax.set_ylabel('Y-axis')
    ax.legend()
    ax.grid(True)

    canvas.draw()  # Redraw the plot
    
def select_profile(event):
    selected_profile_name = combo_var.get()
    optimal_profile_name = '최적대안'
    selected_profile = profile_dict[selected_profile_name]
    plot_alternatives(optimal_profile, top_10_profiles, selected_profile)
    print_profile(selected_profile,selected_profile_name)
    print_profile(optimal_profile,optimal_profile_name)
    
def create_excel(alternative,name):

    # 엑셀 파일 생성
    wb = openpyxl.Workbook()

    # '종단' 시트 생성
    ws1 = wb.active
    ws1.title = "종단"

    # 데이터 쓰기
    # 헤더 추가
    ws1.append(["측점", "계획고"])
    for station, elevation in alternative['종단']:
        ws1.append([station, elevation])

    # '구조물' 시트 생성
    ws2 = wb.create_sheet(title="교량")
    ws3 = wb.create_sheet(title="터널")
    # 데이터 쓰기
    # 헤더 추가
    ws2.append(["명칭", "시점", "종점","길이"])
    ws3.append(["명칭", "시점", "종점","길이"])
    
    for struct_type, (start_station, end_station) in alternative['구조물위치'].items():
        length = end_station - start_station
        if struct_type.startswith('B'):
            ws2.append([struct_type, start_station, end_station,length])
        else:
            ws3.append([struct_type, start_station, end_station,length])
    
    # '절성고' 시트 생성
    ws4 = wb.create_sheet(title="절성고")

    # 데이터 쓰기
    # 헤더 추가
    ws4.append(["측점", "절성토"])
    
    for station,diff in alternative['절성고']:
        ws4.append([station, diff*-1])
    
    # 엑셀 파일 저장
    dierctory = 'c:\\temp\\'
    filename = name + '.xlsx'
    
    wb.save(dierctory + filename)
    
def print_profile(profile,name):
    print("\n-----Selected Profile Information:-----\n")
    print(f'{name}')
    # Calculate and print score
    station_elv = generate_station_elv(profile)
    cutfill = calculate_cutfill(station_elv, gl)
    score, total_cut_volume, total_fill_volume = calculate_cutfill_volume_and_score(cutfill)
    print(f"Score: {score}")
    
    # Calculate and print construction cost
    struct = cal_profile_structure(cutfill)
    
    construction_cost = cal_profile_money(struct)  # You need to define this function
    print(f"공사비 : {construction_cost:.2f}백만원")
    
    # Calculate and print steepest gradient
    steepest_gradient, steepest_gradient_location = find_steepest_gradient(profile)
    print(f'최급기울기 = {steepest_gradient:.2f} 퍼밀')
    print(f'기울기개소 = {steepest_gradient_location} 개소')

    # Determine and print structures
    struct_station = find_structure_ranges(struct)
    count_struct = count_structures(struct_station)
    
    print(f'구조물 = (교량: {count_struct[0]}개소, 터널: {count_struct[1]}개소)')

def profile_spec(profile):
    alternative = {}
    alternative['종단'] = profile
    alternative['계획고'] = generate_station_elv(alternative['종단']) 
    alternative['절성고'] = calculate_cutfill(alternative['계획고'], gl)
    alternative['구조물'] = cal_profile_structure(alternative['절성고'])
    alternative['구조물위치'] = find_structure_ranges(alternative['구조물'])
    alternative['구조물갯수'] = count_structures(alternative['구조물위치'])
    alternative['공사비'] = cal_profile_money(alternative['구조물'])
    alternative['점수'] = calculate_cutfill_volume_and_score(alternative['절성고'])
    
    return alternative

def toggle_aspect_ratio():
    if aspect_var.get() == 1:
        ax.set_aspect((1000/400), adjustable='datalim')
    else:
        ax.set_aspect('auto', adjustable='box')
    canvas.draw()
    
gl = load_gl()
profiles = []
scores = []
profile_dict = {}
#최대 vip갯수
#vip간 최소거리
min_distance = 1000
max_vip = int(gl[-1][0] / min_distance)
print(max_vip)

#최대종단경사(퍼밀)
LIMIT_GRADE = 25

for _ in range(1000):  # 100개의 대안 생성
    profile = generate_random_profile(max_vip, min_distance, gl)
    adjusted_profile = check_and_adjust_elevation(profile)
    max_grade_profile = find_steepest_gradient(adjusted_profile)
    if max_grade_profile[0] < LIMIT_GRADE:
        profiles.append(adjusted_profile)
        station_elv = generate_station_elv(adjusted_profile)
        cutfill = calculate_cutfill(station_elv, gl)
        score, total_cut_volume, total_fill_volume = calculate_cutfill_volume_and_score(cutfill)
        scores.append(score)
  
# 최적 대안 선택

optimal_index = scores.index(max(scores))
optimal_profile = profiles[optimal_index]

# 상위 10개의 대안 선택
top_10_indices = sorted(range(len(scores)), key=lambda i: scores[i])[:10]
top_10_profiles = [profiles[i] for i in top_10_indices]

# 프로파일 딕셔너리 생성
profile_dict = {f"대안 {i + 1}": profile for i, profile in enumerate(top_10_profiles)}



# GUI 구성
root = tk.Tk()
root.title("대안 선택")

combo_var = tk.StringVar(root)
combo_var.set("대안 1")

combo_box = ttk.Combobox(root, textvariable=combo_var, values=list(profile_dict.keys()))
combo_box.pack()

combo_box.bind("<<ComboboxSelected>>", select_profile)

# Create a variable to store the checkbox state
aspect_var = tk.IntVar(value=0)

# Create the checkbox
aspect_checkbox = tk.Checkbutton(root, text="비율 동일하게", variable=aspect_var, command=toggle_aspect_ratio)
aspect_checkbox.pack()


# Initialize the plot
fig = plt.figure(figsize=(6, 6))
ax = fig.add_subplot()
canvas = FigureCanvasTkAgg(fig, master=root)
canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

# Create a frame for the toolbar
toolbar_frame = tk.Frame(root)
toolbar_frame.pack(side=tk.BOTTOM, fill=tk.X)

# Add the navigation toolbar
toolbar = NavigationToolbar2Tk(canvas, toolbar_frame)
toolbar.update()


# Run the GUI
root.geometry("900x800")

# 처음 플롯 생성
plot_alternatives(optimal_profile, top_10_profiles)

root.mainloop()
