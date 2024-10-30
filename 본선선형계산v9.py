from tkinter import filedialog
import csv
from shapely.geometry import LineString, Point
import math
import pandas as pd
from openpyxl import Workbook, load_workbook
import time
import ezdxf
import numpy as np
import matplotlib.pyplot as plt

def create_dxf(coordinates, linestring, data1, data2, filename='test.dxf', scale=1000):
    doc = ezdxf.new()
    msp = doc.modelspace()

    scale = scale / 1000 #1단위로 변환
    
    #data1 = 선형계산데이터
    #data2 = 좌표데이터

    
    # 각 튜플의 요소들을 뒤집기(토목좌표를 수학좌표로 변환)
    coordinates_math = [(y, x) for x, y in coordinates]
    linestring_math = [(y, x) for x, y in linestring.coords]
    
    # Create a polyline entity
    ippolyline = msp.add_lwpolyline(linestring_math)#ip라인
    polyline = msp.add_lwpolyline(coordinates_math)#선형
    
    
    # Set the color of the polyline to red (color index 1)
    '''#ACI인덱스
    0: BYBLOCK
    1: Red
    2: Yellow
    3: Green
    4: Cyan
    5: Blue
    6: Magenta
    7: White
    8: Gray
    9: BYLAYER
    '''
    
    # Set polyline properties if needed
    polyline.dxf.layer = 'FL'
    ippolyline.dxf.layer = 'IP라인'
    
    # Set the color of the polyline to red
    
    polyline.dxf.color = 1
    ippolyline.dxf.color = 9

    #폴리선굵기
    polyline.dxf.const_width  = scale
    
    text_color_index = 7
    text_height = 3 * scale

    #IP좌표테이블
    '''구성
    IPNO.1
    IA = 7D 9' 10"
    R= 600
    TL =145.15
    CL = 145.24
    X = 24236.26
    Y=2515235
    
    '''
    dimention = 2
    coord_precision = 4
    km_precision = 2
    
    for i, entry in enumerate(data1):
        IP_number = f'IP  NO. {entry[0]}'
        IA = f'IA = {degrees_to_dms2(entry[6])}'
        R= f'R = {entry[8]:.{dimention}f}' # R
        if entry[2] == 'simplecurve':  # Simple curve case
            TL = f'TL = {entry[14][0]:.{dimention}f}' #TL
            CL = f'CL = {entry[14][1]:.{dimention}f}' #CL
        else:
            TL = f'TL = {entry[14][10]:.{dimention}f}' #TL
            CL = f'CL = {entry[14][12]:.{dimention}f}' #CL

        X = entry[1][0]
        Y = entry[1][1]
        
        X_TEXT= f'X = {entry[1][0]:.{coord_precision}f}'
        Y_TEXT = f'Y = {entry[1][1]:.{coord_precision}f}'

        '''
        #텍스트 좌표
        # Text coordinates
        text_offsets = [0, -10, -20, -30, -40, -50, -60]
        text_labels = [IP_number, IA, R, TL, CL, X_TEXT, Y_TEXT]

        for offset, label in zip(text_offsets, text_labels):
            coord = (Y, X + offset)
            msp.add_text(label, dxfattribs={'insert': coord, 'height': text_height, 'color': text_color_index, 'layer': 'table'})
        '''

        # Create block definition with the text elements
        block_name = f"IP_Block_{i}"
        block = doc.blocks.new(name=block_name)
        
        text_labels = [IP_number, IA, R, TL, CL, X_TEXT, Y_TEXT]
        text_offsets = [0 * scale, -10 * scale, -20 * scale, -30 * scale, -40 * scale, -50 * scale, -60 * scale]
        
        for offset, label in zip(text_offsets, text_labels):
            coord = (entry[1][1], entry[1][0] + offset)
            block.add_text(label, dxfattribs={
                'insert': coord, 
                'height': text_height, 
                'color': text_color_index,
                'layer': 'IP제원표'
            })
        
        # Insert block into the model space at IP location
        msp.add_blockref(block_name, insert=(0, 0))
        
        #선형제원문자
        # Write the relevant station data for the current entry
        if i < len(data2):  # Ensure you do not exceed data2 bounds
            station_group = data2[i]  # Get the corresponding station group for the current entry
            for station_data in station_group:
                if i != 0:
                    station_data.pop(0)
                
                
                for row in station_data:  # Iterate through the innermost list
                    
                    lable, sta, station_offset, coord, shiftx, shifty , Ta, sageoriS , Q, T, Azimuth = row

                    #토목좌표를 수학좌표로 변환
                    x, y = coord
                    coord_math = (y ,x)

                    #접선방위각을 수학각도로 변환
                    Ta_math = (450 - Ta) % 360
                    
                    #km문자
                    if sta % 1000 == 0 or sta == 0:#1000의 배수인경우(0포함)
                        #원그리기
                        msp.add_circle(coord_math, radius = 2 * scale, dxfattribs={'layer': 'km심볼', 'color': 1})
                        #원그리기
                        msp.add_circle(coord_math, radius = 1.5 * scale, dxfattribs={'layer': '200심볼', 'color' : 1})
                        
                        #문자
                        kmtext = f'{sta // 1000}km'
                        msp.add_text(kmtext, dxfattribs={'layer': 'km문자', 'insert': coord_math, 'height': text_height, 'color': 1, 'rotation': Ta_math})

                        #선그리기
                        #시작점과 끝점 계산
                        start_point = calculate_destination_coordinates(y, x, Ta_math + 90, 1.5 * scale)
                        end_point = calculate_destination_coordinates(y, x, Ta_math + 90, -1.5 * scale)
                        msp.add_line(start_point, end_point, dxfattribs={'layer': '보조측점', 'color': 1})#중심점,길이
                        
                    #200문자
                    elif sta % 200 == 0:#200의 배수인경우(0포함)
                        #원그리기
                        msp.add_circle(coord_math, radius = 1.5 * scale, dxfattribs={'layer': '200심볼', 'color' : 1})
                        #200 텍스트
                        text200 = str(sta % 1000)
                        msp.add_text(text200, dxfattribs={'layer': '200문자', 'insert': coord_math, 'height': text_height, 'color': 1, 'rotation': Ta_math})

                        #선그리기
                        #시작점과 끝점 계산
                        start_point = calculate_destination_coordinates(y, x, Ta_math + 90, 1.5 * scale)
                        end_point = calculate_destination_coordinates(y, x, Ta_math + 90, -1.5 * scale)
                        msp.add_line(start_point, end_point, dxfattribs={'layer': '보조측점', 'color': 1})#중심점,길이
                        
                    #보조측점선
                    elif sta % unit == 0:
                        
                        #선그리기
                        #시작점과 끝점 계산
                        start_point = calculate_destination_coordinates(y, x, Ta_math + 90, 1.5 * scale)
                        end_point = calculate_destination_coordinates(y, x, Ta_math + 90, -1.5 * scale)
                        msp.add_line(start_point, end_point, dxfattribs={'layer': '보조측점', 'color': 1})#중심점,길이


                        
                    #완화곡선인출선
                    else:
                        
                        #시작점과 끝점 계산
                        start_point = coord_math
                        end_point = calculate_destination_coordinates(y, x, Ta_math + 90, 50 * scale) if entry[3] == 'LEFT CURVE' else calculate_destination_coordinates(y, x, Ta_math - 90, 50 * scale)
                        msp.add_line(start_point, end_point, dxfattribs={'layer': '곡선인출선', 'color': 1})#중심점,길이

                        #제원문자
                        text_rotation = Ta_math - 90
                        reverse_rotaion = Ta_math + 90 + 180 if entry[3] == 'LEFT CURVE' else Ta_math - 90 + 180
                        
                        if sta / 1000 < 10:
                            text_position = calculate_destination_coordinates(y, x, reverse_rotaion, -20 * scale)
                        elif sta / 1000 < 100:
                            text_position = calculate_destination_coordinates(y, x, reverse_rotaion, -16 * scale)
                        else:
                            text_position = calculate_destination_coordinates(y, x, reverse_rotaion, -12 * scale)
                            
                        if entry[3] == 'LEFT CURVE':
                            msp.add_text(lable + '=' + str(format_distance(sta, decimal_places=km_precision)), dxfattribs={'layer': '곡선제원문자', 'insert': end_point, 'height': text_height, 'color': 1, 'rotation': text_rotation})
                        else:
                            msp.add_text(lable + '=' + str(format_distance(sta, decimal_places=km_precision)), dxfattribs={'layer': '곡선제원문자', 'insert': text_position, 'height': text_height, 'color': 1, 'rotation': text_rotation})

                     #정거장중심
                     #todo
    #IP좌표 선
    for i in range(len(linestring_math)-2):

        x1, y1  = linestring_math[i]
        x2, y2 = linestring_math[i+1]
        x3 , y3 = linestring_math[i+2]
        
        #v1 계산
        v1 = calculate_bearing(x1, y1, x2, y2)
        v2 = calculate_bearing(x2, y2, x3, y3)
        
        #v1 시작점 계산
        start_point1 = calculate_destination_coordinates(x2 ,y2 , v1 +180, 7 * scale)
        end_point1 = calculate_destination_coordinates(x2 ,y2,  v1, 3 * scale)

        msp.add_line(start_point1, end_point1, dxfattribs={'layer': 'IP교점선', 'color': 1})#중심점,길이
        
        #v2 시작점 계산
        start_point2 = (x2, y2)
        end_point2 = calculate_destination_coordinates(x2 ,y2 ,  v2, 7* scale)
        msp.add_line(start_point2, end_point2, dxfattribs={'layer': 'IP교점선', 'color': 1})#중심점,길이

        #문자
        msp.add_text(f'IP NO.{i+1}' , dxfattribs={
            'layer': 'IP문자',
            'insert': start_point2,
            'height': text_height,
            'color': 1,
            'rotation': 0
            })

        
    # Save the DXF document to a file
    dxf_save_with_retry(doc, filename, max_retries=100, delay=1)
    

def dxf_save_with_retry(doc, filename, max_retries=100, delay=1):
    # Try saving the file, retrying if there's a PermissionError
    for attempt in range(max_retries):
        try:
            doc.saveas("C:/temp/" + filename)
            print(f"File saved successfully: {filename}")
            break
        except PermissionError:
            print(f"Attempt {attempt + 1}: Permission denied while saving {filename}. Retrying in {delay} seconds...")
            time.sleep(delay)
    else:
        print(f"Failed to save the file after {max_retries} attempts.")

            
def save_with_retry(workbook, filename, max_retries=100, delay=1):
    # Try saving the file, retrying if there's a PermissionError
    for attempt in range(max_retries):
        try:
            workbook.save(filename)
            print(f"File saved successfully: {filename}")
            break
        except PermissionError:
            print(f"Attempt {attempt + 1}: Permission denied while saving {filename}. Retrying in {delay} seconds...")
            time.sleep(delay)
    else:
        print(f"Failed to save the file after {max_retries} attempts.")
        
def pasing_ip(lines):
    alignment = []
    for line in lines:
        try:
            IP_X_Coordinate = float(line[1])
            IP_Y_Coordinate = float(line[0])
            Radius = float(line[2])
            
        except IndexError:
            Radius = 0
        
        alignment.append([IP_X_Coordinate, IP_Y_Coordinate, Radius])
        
    return alignment

def create_Linestring(alignment):
    coords = [(point[0], point[1]) for point in alignment]
    return LineString(coords)

def read_file():
    file_path = filedialog.askopenfilename(defaultextension=".txt", filetypes=[("txt files", "*.txt"), ("All files", "*.*")])
    print('현재파일:', file_path)
    
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            lines = list(reader)
    except UnicodeDecodeError:
        print('현재파일은 utf-8인코딩이 아닙니다. euc-kr로 시도합니다.')
        try:
            with open(file_path, 'r', encoding='euc-kr') as file:
                reader = csv.reader(file)
                lines = list(reader)
        except UnicodeDecodeError:
            print('현재파일은 euc-kr인코딩이 아닙니다. 파일을 읽을 수 없습니다.')
            return []
    return lines

def calculate_angles_and_plot(line_string):
    angles = []
    for i in range(1, len(line_string.coords) - 1):
        prev_vector = (line_string.coords[i][0] - line_string.coords[i - 1][0],
                       line_string.coords[i][1] - line_string.coords[i - 1][1])
        current_vector = (line_string.coords[i + 1][0] - line_string.coords[i][0],
                          line_string.coords[i + 1][1] - line_string.coords[i][1])
        angle_diff = calculate_inner_angle(prev_vector, current_vector)
        angles.append(angle_diff)
    return angles

def calculate_inner_angle(v1, v2):
    angle1 = math.atan2(v1[1], v1[0])
    angle2 = math.atan2(v2[1], v2[0])
    angle_diff = abs(angle2 - angle1)
    if angle_diff > math.pi:
        angle_diff = 2 * math.pi - angle_diff
    return math.degrees(angle_diff)

def calculate_bearing(x1, y1, x2, y2):
    dx = x2 - x1
    dy = y2 - y1
    bearing = math.degrees(math.atan2(dy, dx))
    return bearing

def calculate_bearingN(x1, y1, x2, y2):
    # 방위각 계산함수
    dx = x2 - x1
    dy = y2 - y1
    bearing = math.degrees(math.atan2(dy, dx))
    if bearing <=0:
        bearing += 360
    return bearing

def calculate_distance(x1, y1, x2, y2):
    return math.sqrt((x2 - x1)**2 + (y2 - y1)**2)

def calculate_destination_coordinates(x1, y1, bearing, distance):
    angle = math.radians(bearing)
    x2 = x1 + distance * math.cos(angle)
    y2 = y1 + distance * math.sin(angle)
    return x2, y2

def find_direction(d10, v10, v11):
    if d10 == 0:
        return 0
    elif math.sin(math.radians(v11 - v10)) >= 0:
        return 1
    else:
        return -1
    
def calculate_V(Cm, Cd, R, maxV):
    V = int(math.sqrt((Cm +  Cd) * R / 11.8))  # 정수로 변환
    return min(V, maxV)  # V가 maxV를 초과하지 않도록 제한

#최대 설정캔트와 최대 부족캔트
def get_cant_limits(V, track_type):
    """
    설계속도(V)와 도상 유형(track_type)에 따른 최대 설정캔트와 최대 부족캔트를 반환하는 함수.
    
    track_type: '자갈도상' 또는 '콘크리트도상'
    """
    if track_type == "자갈도상":
        if 350 < V <= 400:
            return 180, 130#이 설계속도에서는 콘크리트도상 적용
        elif 250 < V <= 350:
            return 160, 80
        elif V <= 250:
            return 160, 100
    elif track_type == "콘크리트도상":
        if 350 < V <= 400:
            return 180, 130
        elif 250 < V <= 350:
            return 180, 130
        elif V <= 250:
            return 180, 130
    else:
        return "Invalid track type"

#M,Z,V계산
def cal_parameter(radius_list, maxV=250, track_type='자갈도상'):
    parameters = []
    for i in range(len(radius_list)):
        try:
            
            
            Cm, Cd = get_cant_limits(maxV, track_type)#최대 설정캔트와 최대 부족캔트
            
            R = radius_list[i]
            
            
            V = int(calculate_V(Cm, Cd, R, maxV))
            
            # 10의 배수로 조정
            V = round(V / 10) * 10
            
            m_values = {
            70: 600,
            120: 900,
            150: 1100,
            200: 1450,
            250: 1850,
            300: 2200,
            350: 2550,
            400: 2950
            }

            M = int(m_values.get(V, 7.31 * V))
            # 50의 배수로 조정
            M = round(M / 50) * 50
            
            Z = 11.8 * V ** 2 / R
            
            if Z >= Cm:
                Z = Cm

            # 10의 배수로 조정
            Z = round(Z / 10) * 10
            
            parameters.append([M,Z,V])
        except ValueError:
            print("잘못된 입력입니다. 숫자를 입력해주세요.")
            
    return parameters

def degrees_to_dms(degrees):
    """
    Converts decimal degrees to degrees, minutes, seconds.
    
    Args:
    degrees (float): Decimal degrees value.
    
    Returns:
    tuple: Degrees, minutes, seconds.
    """
    if degrees < 0:
        degrees = degrees * -1
        
    deg = int(degrees)
    minutes = int((degrees - deg) * 60)
    seconds = (degrees - deg - minutes / 60) * 3600

    
    return f"{deg}도  {minutes}분  {seconds:.2f}초"

def degrees_to_dms2(degrees):
    """
    Converts decimal degrees to degrees, minutes, seconds.
    
    Args:
    degrees (float): Decimal degrees value.
    
    Returns:
    tuple: Degrees, minutes, seconds.
    """
    if degrees < 0:
        degrees = degrees * -1
        
    deg = int(degrees)
    minutes = int((degrees - deg) * 60)
    seconds = (degrees - deg - minutes / 60) * 3600

    
    return f"{deg}° {minutes}' {seconds:.2f}\""

#SP>PC 방위각
def calculate_SP_PC_bearing(W13, V10, X10):
    if W13 < 0:
        result = V10 - X10
        if result < 0:
            result += 360
    else:
        result = V10 + X10
        if result > 360:
            result -= 360

    return result

#CP좌표계산
def calculate_CP_XY(K21, N21, W13, W18, P10, V11):
    part1 = W13 * math.cos(math.radians(W18 + 90))  # (W18 + 90) * PI() / 180
    part2 = W13 * math.cos(math.radians(W18 - 90 + (180 * P10) / (math.pi * W13)))  # (W18 - 90 + (180 * P10 / PI() / W13)) * PI() / 180
    part3 = W13 * math.sin(math.radians(W18 + 90))
    part4 = W13 * math.sin(math.radians(W18 - 90 + (180 * P10) / (math.pi * W13)))  # (W18 - 90 + (180 * P10 / PI() / W13)) * PI() / 180
    
    result1 = K21 + part1 + part2
    result2 = N21 + part3 + part4
    
    return (result1, result2)


#O>PC방위각
def calculate_O_PC_bearing(V10, X10, W13):
    '''args
    W13 = W13
    V10 = 방위각1
    X10 = THITA
    '''
    X10 = math.degrees(X10)
    
    if W13 > 0:
        if V10 + X10 - 90 < 0:
            return V10 + X10 - 90 + 360
        else:
            return V10 + X10 - 90
    else:
        if V10 - X10 + 90 > 360:
            return V10 - X10 + 90 - 360
        else:
            return V10 - X10 + 90

#PC점의 접선각
def calculate_theta_pc(m, z, Rc):
    """Calculate the PC point's tangent angle in radians."""
    x1 = m * (z * 0.001) #X1
    theta_pc = math.atan(x1 / (2 * Rc))
    return theta_pc

#3차포물선 제원계산
def calculate_cubic_parabola_parameter(m, z, direction, theta_pc, Rc, IA_rad):
    """Calculate various lengths used in the spiral calculation."""
    x1 = m * (z * 0.001) #X1
    x2 = x1 - (Rc * math.sin(theta_pc))#x2
    W13 = Rc * direction#-R
    L = x1 * (1 + (math.tan(theta_pc) ** 2) / 10)#완화곡선 길이
    Y = (x1 ** 2) / (6 * Rc)#Y1
    W15  = (x1 ** 2) / (6 * W13)#W-Y
    F = Y - Rc * (1 - math.cos(theta_pc))#이정량 F
    S = 1 / math.cos(IA_rad / 2) * F #S
    K = F * math.tan(IA_rad / 2) #수평좌표차K
    W = (Rc + F) * math.tan(IA_rad / 2)#W
    TL = x2 + W #TL
    Lc = Rc * (IA_rad - 2 * theta_pc) #원곡선 길이
    CL = Lc + 2 * L #전체곡선길이
    SL = Rc * (1 / (math.cos(IA_rad / 2)) - 1) + S
    RIA = IA_rad - (2 * theta_pc)#원곡선교각
    C = math.atan(Y / x1)#C
    XB = math.pi/2 - theta_pc#XB
    B = math.pi/2 - C - XB
    parameter = [x1, x2, W13, L, Y, W15, F, S, K, W, TL, Lc, CL, SL, RIA , C, B]
    
    return parameter

def calculate_simplecurve_parameter(direction, Rc, IA_rad):
    TL = Rc * math.tan(IA_rad/2)
    CL = Rc * IA_rad
    SL = Rc * (1 / (math.cos(IA_rad / 2)) - 1)
    S = Rc * (1 - math.cos(IA_rad / 2))
    W13 = Rc * direction
    parameter = [TL,CL,SL,S, W13]
    return parameter

#완화곡선 시종점 좌표계산함수
def calculate_spiral_coordinates(BP_XY, IP_XY, EP_XY, h1, h2, TL, x1, Y, SP_PC_bearing, Rc, Lc, W13):
    """Calculate the coordinates of SP, PC, CP, and PS."""
    SP_XY = (
        math.cos(math.radians(h1 + 180)) * TL + IP_XY[0],
        math.sin(math.radians(h1 + 180)) * TL + IP_XY[1]
    )
    PC_XY = (
        SP_XY[0] + x1 * math.cos(math.radians(h1)) + Y * math.cos(math.radians(h1 + 90)),
        SP_XY[1] + x1 * math.sin(math.radians(h1)) + Y * math.sin(math.radians(h1 + 90))
    )
    CP_XY = calculate_CP_XY(PC_XY[0], PC_XY[1], W13, SP_PC_bearing, Lc, h2)
    PS_XY = (
        math.cos(math.radians(h2)) * TL + IP_XY[0],
        math.sin(math.radians(h2)) * TL + IP_XY[1]
    )
    return SP_XY, PC_XY, CP_XY, PS_XY

def dictionary_cubicparabola(v):
    '''
    완화곡선을 삽입하지 않는 최소곡선반경
    '''
    not_spiral_min_radius_dic = {
        250: 24000,
        200: 12000,
        150: 5000,
        120: 2500,
        70: 600
    }

    cdlim_dic = {
        400: 20,
        350: 25,
        300: 27,
        250: 32,
        200: 40,
        150: 57,
        120: 69,
        100: 83,
        70: 100
    }

    # cdlim_dic에서 이미 존재하는 값일 경우 그대로 반환
    if v in cdlim_dic:
        cdlim = cdlim_dic[v]
    else:
        # 값이 사전에 없을 경우 선형 보간 수행
        sorted_keys = sorted(cdlim_dic.keys())

        # 입력값이 범위 밖에 있을 경우 예외 처리
        if v < sorted_keys[0] or v > sorted_keys[-1]:
            raise ValueError(f"입력 속도 {v}는 지원되는 범위 밖에 있습니다. {sorted_keys[0]}에서 {sorted_keys[-1]} 사이의 값을 입력하세요.")

        # 선형 보간을 위한 인접한 두 값 찾기
        for i in range(len(sorted_keys) - 1):
            if sorted_keys[i] <= v <= sorted_keys[i + 1]:
                # 두 지점 사이의 보간
                x1, y1 = sorted_keys[i], cdlim_dic[sorted_keys[i]]
                x2, y2 = sorted_keys[i + 1], cdlim_dic[sorted_keys[i + 1]]
                
                # 선형 보간 공식 사용
                cdlim = y1 + (v - x1) * (y2 - y1) / (x2 - x1)
                break

    # not_spiral_min_radius_dic에서 값 찾기
    if v in not_spiral_min_radius_dic:
        not_spiral_min_radius = not_spiral_min_radius_dic[v]
    else:
        # 해당 값이 없으면 공식으로 계산
        not_spiral_min_radius = 11.8 * (v**2 / cdlim)

    return not_spiral_min_radius


#메인계산함수
def calculate_curve(linestring, Radius_list, angles, parameters, unit=20, start_STA=0):

    final_result = []
    
    O_XY_list = []
    cubic_parabola_XY_list = []
    cubic_parabola_STA_list = []

    simplecurve_XY_list = []
    simplecurve_STA_list = []
    
    TL_LIST = []
    IP_STA_list = []
    
    direction_list = []
    prebearing_list = []
    nextbearing_list = []
    
    BP_STA = start_STA
    BP_STA_LIST = [0]

    station_coordlist = []

    END_STA_list = []
    END_XY_list =[]

    alignment_report_variable_list = []

    polycurve = []
    for i in range(len(linestring.coords)-2):
        IA_rad = math.radians(angles[i])
        IA_DEGREE = angles[i]
        
        #토목좌표
        BP_XY = (linestring.coords[i][0], linestring.coords[i][1])
        IP_XY = (linestring.coords[i + 1][0], linestring.coords[i + 1][1])
        EP_XY = (linestring.coords[i + 2][0], linestring.coords[i + 2][1])

        h1 = calculate_bearingN(BP_XY[0], BP_XY[1], IP_XY[0], IP_XY[1])#방위각1(도)
        h2 = calculate_bearingN(IP_XY[0], IP_XY[1], EP_XY[0], EP_XY[1])#방위각2(도)

        IP_distnace1 = calculate_distance(BP_XY[0], BP_XY[1], IP_XY[0], IP_XY[1]) #ip연장1
        IP_distnace2 = calculate_distance(IP_XY[0], IP_XY[1], EP_XY[0], EP_XY[1]) #연장2

        
        
        Rc = Radius_list[i]
        m, z, v = parameters[i]
        
        not_spiral_min_radius = dictionary_cubicparabola(v)#완화곡선을 삽입하지 않는 최소곡선반경

        if Rc >= not_spiral_min_radius:
            curvetype = 'simplecurve'#단곡선
        else:
            curvetype = 'cubic'#3차포물선
        
        direction = find_direction(Rc, h1, h2)

        if direction == 1:
            curvedirection = 'RIGHT CURVE'
        else:
            curvedirection = 'LEFT CURVE'

        #마지막 IP변수
        is_last_ip = (i == len(linestring.coords) - 3 if len(linestring.coords) > 3 else i == len(linestring.coords) - 2)
        
        if curvetype == 'simplecurve':#단곡선
            theta_pc = 0
            #O>PC 방위각(도)
            
            simplecurve_parameter = calculate_simplecurve_parameter(direction, Rc, IA_rad)
            TL, CL, SL, S, W13 = simplecurve_parameter# 파라매터

            O_PC_bearing = calculate_O_PC_bearing(h1, theta_pc, W13)
            SP_PC_bearing = h1#SP>PC방위각(도)
            simplecurve_coord = calculate_simplecurve_coordinates(h1, h2, TL, IP_XY)#BC,EC좌표반환함수
            BC_XY , EC_XY = simplecurve_coord#BC,EC좌표
            
            CURVE_CENTER = (math.cos(math.radians(O_PC_bearing + 180)) * Rc + BC_XY[0],
                            math.sin(math.radians(O_PC_bearing + 180)) * Rc + BC_XY[1])#CURVE CENTER

            simplecurve_XY_list.append([BC_XY, EC_XY])

            END_XY_list.append(EC_XY)
            
            if i == 0:
                simplecurve_STA_list.append([0, 0])#인덱스 오류 방지로 0 할당
                # 첫 번째 값을 덮어쓰기
                IP_STA = BP_STA + IP_distnace1
                BC_STA = IP_STA - TL
                EC_STA = BC_STA + CL

                simplecurve_STA_list[0] = [BC_STA, EC_STA]  # 첫 번째 값 덮어쓰기

                END_STA_list.append(EC_STA)
            else:
                BP_STA = END_STA_list[i-1]  # 이전 PS_STA 값
                IP_STA = BP_STA + IP_distnace1 - TL_LIST[i - 1]

                BC_STA = IP_STA - TL
                EC_STA = BC_STA + CL

                simplecurve_STA_list.append([BC_STA, EC_STA])
                
                END_STA_list.append(EC_STA)
                if is_last_ip:#마지막IP
                    EC_EP_distance = calculate_distance(EC_XY[0], EC_XY[1], EP_XY[0], EP_XY[1]) #EC와 EP거리
                    EP_STA = EC_STA + EC_EP_distance
        else:#3차포물선
            theta_pc = calculate_theta_pc(m, z, Rc)#pc점의 접선각 라디안
            
            
            cubic_parameter = calculate_cubic_parabola_parameter(m, z, direction, theta_pc, Rc, IA_rad)#파라매터
        
            x1, x2, W13, L, Y, W15, F, S, K, W, TL, Lc, CL, SL, RIA, C, B = cubic_parameter

            O_PC_bearing = calculate_O_PC_bearing(h1, theta_pc, W13)#O>PC 방위각(도)
            
            SP_PC_bearing = calculate_SP_PC_bearing(W13, h1, math.degrees(theta_pc))#SP>PC방위각(도)

            SP_XY, PC_XY, CP_XY, PS_XY = calculate_spiral_coordinates(BP_XY, IP_XY, EP_XY, h1, h2, TL, x1, W15, SP_PC_bearing, Rc, Lc, W13)

            CURVE_CENTER = (math.cos(math.radians(O_PC_bearing + 180)) * Rc + PC_XY[0],
                            math.sin(math.radians(O_PC_bearing + 180)) * Rc + PC_XY[1])#CURVE CENTER
            
            cubic_parabola_XY_list.append([SP_XY,PC_XY,CP_XY,PS_XY])

            END_XY_list.append(PS_XY)
            
            if i == 0:
                cubic_parabola_STA_list.append([0, 0, 0, 0])#인덱스 오류 방지로 0 할당
                # 첫 번째 값을 덮어쓰기
                IP_STA = BP_STA + IP_distnace1
                SP_STA = IP_STA - TL
                PC_STA = SP_STA + L
                CP_STA = PC_STA + Lc
                PS_STA = CP_STA + L

                cubic_parabola_STA_list[0] = [SP_STA, PC_STA, CP_STA, PS_STA]  # 첫 번째 값 덮어쓰기

                END_STA_list.append(PS_STA)
                
            else:
                # 두 번째 이상의 경우 새로운 값을 append
                BP_STA = END_STA_list[i-1]  # 이전 PS_STA 값
                IP_STA = BP_STA + IP_distnace1 - TL_LIST[i - 1]

                SP_STA = IP_STA - TL
                PC_STA = SP_STA + L
                CP_STA = PC_STA + Lc
                PS_STA = CP_STA + L

                cubic_parabola_STA_list.append([SP_STA, PC_STA, CP_STA, PS_STA])
                
                END_STA_list.append(PS_STA)

                if is_last_ip:#마지막IP
                    EC_EP_distance = calculate_distance(PS_XY[0], PS_XY[1], EP_XY[0], EP_XY[1]) #EC와 EP거리
                    EP_STA = PS_STA + EC_EP_distance

            #경고추가
            if BP_STA>SP_STA:
                print(f'경고! 곡선겹침 IP{i+1}: BP={format_distance(BP_STA)}, SP={format_distance(SP_STA)}')
                
        TL_LIST.append(TL)
        
        
            
        
            
        ###좌표 계산로직###
        # unit 간격으로 PS_STA까지 스테이션 리스트 생성
        
        #리스트 초기화
        station_coordlist = []
        
        # Ensure BP_STA and PS_STA are integers

        
        
        if curvetype == 'simplecurve':#단곡선
            BP_STA_int = int(round(BP_STA))
            EC_STA_int = int(round(EC_STA))


            # unit 간격으로 PS_STA까지 스테이션 리스트 생성
            if is_last_ip: #마지막IP
                EP_STA_int = int(round(EP_STA))
                station_list = generate_integers(BP_STA_int, EP_STA_int, unit)
            else:
                station_list = generate_integers(BP_STA_int, EC_STA_int, unit)
                
            #BC와 EC측점 추가
            station_list.extend([BP_STA, BC_STA, EC_STA])

            #리스트 정렬
            station_list.sort()

            # 스테이션 리스트에 있는 각 sta에 대해 좌표 계산
            for j, sta in enumerate(station_list):
                if j ==0:
                    lable = 'BP'
                    StationOffset = 0
                else:
                    if is_last_ip:#마지막 IP
                        lable = get_station_label(curvetype, sta, BC_STA, BC_STA, EC_STA, EC_STA, EP_STA)  # 제원문자
                    else:
                        lable = get_station_label(curvetype, sta, BC_STA, BC_STA, EC_STA, EC_STA, EC_STA)#제원문자

                    StationOffset = calculate_station_offset(sta, BC_STA, BC_STA, EC_STA, EC_STA, BP_STA)#측거
                #print(StationOffset)
                shiftx =  0
                #print(shiftx)
                shifty =  0
                #print(shifty)
                if i ==0:
                    PBP_XY = BP_XY
                else:
                    PBP_XY = END_XY_list[i-1]
                coord = calculate_coordinates(sta,
                                              BC_STA, BC_STA, EC_STA, EC_STA,
                                              h1, h2, SP_PC_bearing,
                                              StationOffset, W13,
                                              shiftx, shifty,
                                              PBP_XY[0], PBP_XY[1],
                                              BC_XY[0], BC_XY[1],
                                              CURVE_CENTER[0], CURVE_CENTER[1],
                                              EC_XY[0], EC_XY[1])
                # 각 sta에 대한 좌표를 추가

                
                if BC_STA < sta <= EC_STA:#단곡선구간
                    #사거리S
                    sageoriS = 2 * Rc * math.sin((((StationOffset / CL) * IA_rad) / 2) * math.pi / 180)
                    
                    #편기각 Q                   
                    Q = (180 / (2 * math.pi)) * (StationOffset / Rc)#도

                    #접선각 T
                    T = math.degrees(theta_pc) #도
                    
                    # 방위각 A(도) (Azimuth in degrees)
                    Azimuth = h1 + math.degrees(theta_pc) + Q if direction == 1 else h1 - math.degrees(
                        theta_pc) - Q


                    # 접선방위각 Ta (Tangent Azimuth in degrees)
                    Ta = h1 + math.degrees(theta_pc) + 2 * Q if direction == 1 else h1 - math.degrees(
                        theta_pc) - 2 * Q

                elif sta <= BC_STA:#BP - SP
                    sageoriS = 0
                    Ta = h1
                    Q =0
                    T = 0
                    Azimuth = 0
                    
                elif EC_STA <= sta: #PS>EP
                    sageoriS = 0
                    Ta = h2
                    Q =0
                    T = 0
                    Azimuth = 0
                    
                station_coordlist.append([lable, sta, StationOffset, coord, shiftx, shifty, Ta, sageoriS , Q, T, Azimuth])
                
                polycurve.append(coord)
        else:#3차포물선
            BP_STA_int = int(round(BP_STA))
            PS_STA_int = int(round(PS_STA))


            # unit 간격으로 PS_STA까지 스테이션 리스트 생성
            if is_last_ip: #마지막IP
                
                EP_STA_int = int(round(EP_STA))
                station_list = generate_integers(BP_STA_int, EP_STA_int, unit)
            else:

                station_list = generate_integers(BP_STA_int, PS_STA_int, unit)
            
            #BP, SP,PC,CP ,PS점 추가
            if BP_STA % unit == 0:#배수인경우 BP_STA 추가안함
                station_list.extend([SP_STA, PC_STA, CP_STA, PS_STA])
            else:
                station_list.extend([BP_STA, SP_STA, PC_STA, CP_STA, PS_STA])
            
            if is_last_ip: #마지막IP
                station_list.append(EP_STA)
            
            #리스트 정렬
            station_list.sort()

            # 스테이션 리스트에 있는 각 sta에 대해 좌표 계산
            for j, sta in enumerate(station_list):
                if j == 0:
                    lable = 'BP'
                    StationOffset = 0
                else:
                    end_sta = EP_STA if is_last_ip else PS_STA
                    lable = get_station_label(curvetype, sta, SP_STA, PC_STA, CP_STA, PS_STA, end_sta)  # 제원문자

                    StationOffset = calculate_station_offset(sta, SP_STA, PC_STA, CP_STA, PS_STA, BP_STA)#측거

                #이정량X
                shiftx =  cal_spiral_shiftx(sta, SP_STA, PC_STA, CP_STA, PS_STA, StationOffset, W13, L)
                #이정량Y
                shifty =  cal_spiral_shifty(sta, SP_STA, PC_STA, CP_STA, PS_STA, shiftx, W13, x1)

                #사거리 S ageori
                sageoriS = math.sqrt(shiftx**2 + shifty**2)
                
                
                if PC_STA <= sta <= CP_STA:#단곡선구간
                    
                    #사거리S
                    sageoriS = 2 * Rc * math.sin((((StationOffset / Lc) * RIA) / 2) * math.pi / 180)
                    
                    #편기각 Q                   
                    Q = (180 / (2 * math.pi)) * (StationOffset / Rc)#도

                    #접선각 T
                    T = math.degrees(theta_pc) #도
                    
                    # 방위각 A(도) (Azimuth in degrees)
                    Azimuth = h1 + math.degrees(theta_pc) + Q if direction == 1 else h1 - math.degrees(
                        theta_pc) - Q


                    # 접선방위각 Ta (Tangent Azimuth in degrees)
                    Ta = h1 + math.degrees(theta_pc) + 2 * Q if direction == 1 else h1 - math.degrees(
                        theta_pc) - 2 * Q

                    
                    
                elif SP_STA < sta < PC_STA:#SP>PC
                    #편기각 Q
                    Q = math.atan(shiftx**2 / (6 * Rc * x1)) #type 라디안
                    
                    #접선각 Tangent
                    T = (180*shiftx**2) / (6 * math.pi * Rc * L)*3 #type 도
                
                    #방위각 Azimuth
                
                    Azimuth = h1 + math.degrees(Q) if direction == 1 else h1 - math.degrees(Q)#도
                
                    #접선방위각 Ta
                    Ta = h1 + T if direction == 1 else h1 - T
                    
                elif CP_STA < sta < PS_STA:#CP>PS
                    
                    #편기각 Q
                    Q = math.atan(shiftx**2 / (6 * Rc * x1))#type 라디안
                    
                    #접선각 Tangent
                    T = (180*shiftx**2) / (6 * math.pi * Rc * L)*3 #type 도
                
                    #방위각 Azimuth
                
                    Azimuth = h2 - math.degrees(Q) if direction == 1 else h2 + math.degrees(Q)#도
                
                    #접선방위각 Ta
                    Ta = h2 - T if direction == 1 else h2 + T
                    
                elif sta <= SP_STA:#BP - SP
                    sageoriS = 0
                    Ta = h1
                    Q =0
                    T = 0
                    Azimuth = 0
                    
                elif PS_STA <= sta: #PS>EP
                    sageoriS = 0
                    Ta = h2
                    Q =0
                    T = 0
                    Azimuth = 0
                        
                if i ==0:
                    PBP_XY = BP_XY
                else:
                    PBP_XY = END_XY_list[i-1]
                coord = calculate_coordinates(sta,
                                              SP_STA, PC_STA, CP_STA, PS_STA,
                                              h1, h2, SP_PC_bearing,
                                              StationOffset, W13,
                                              shiftx, shifty,
                                              PBP_XY[0], PBP_XY[1],
                                              SP_XY[0], SP_XY[1],
                                              CURVE_CENTER[0], CURVE_CENTER[1],
                                              PS_XY[0], PS_XY[1])

                # 각 sta에 대한 좌표를 추가
                # Create a list for the current station's data
                station_coordlist.append([lable, sta, StationOffset, coord, shiftx, shifty, Ta, sageoriS , Q, T, Azimuth])
                polycurve.append(coord)
        final_result.append([station_coordlist])

        
        ######선형계산서 리스트 저장#######
        if curvetype == 'simplecurve':#단곡선
            
            alignment_report_variable_list.append(
                [
                    i+1,
                    IP_XY,
                    curvetype,
                    curvedirection,
                    h1, h2,
                    IA_DEGREE,
                    CURVE_CENTER,
                    Rc,
                    IP_distnace1, IP_distnace2,
                    m, z, v,
                    simplecurve_parameter,
                    O_PC_bearing,
                    SP_PC_bearing,
                    theta_pc,
                    IP_STA,
                    BC_STA, EC_STA,  # Corrected typo here
                    BC_XY, EC_XY
                ]
            )
        else:#3차포물선
            alignment_report_variable_list.append(
                [
                    i+1,
                    IP_XY,
                    curvetype,
                    curvedirection,
                    h1, h2,
                    IA_DEGREE,
                    CURVE_CENTER,
                    Rc,
                    IP_distnace1, IP_distnace2,
                    m, z, v,
                    cubic_parameter,
                    O_PC_bearing,
                    SP_PC_bearing,
                    theta_pc,
                    IP_STA,
                    SP_STA, PC_STA, CP_STA, PS_STA,# Corrected typo here
                    SP_XY, PC_XY ,CP_XY, PS_XY
                ]
            )
            
                                            
    return final_result, alignment_report_variable_list ,polycurve

def calculate_simplecurve_coordinates(V10, V11, D12, IP):
    '''
    Args:
    v10 = 방위각1
    V11 = 방위각2
    d12 = TL
    F5 =IPX
    L5 = IPY
    
    '''
    F5 = IP[0]
    L5 = IP[1]
    #COS((V10+180)*PI()/180)*D12+F5)
    BC_XY = (math.cos((V10 + 180) * math.pi/180) * D12 + F5,
             math.sin((V10 + 180) * math.pi/180) * D12 + L5)
    EC_XY = (math.cos((V11) * math.pi/180) * D12 + F5,
             math.sin((V11) * math.pi/180) * D12 + L5)

    coords = [BC_XY, EC_XY]
    return coords
#이정량 x,y                
def cal_spiral_shiftx(B27, H20, H21, H22, H23, D27, W13, D13, epsilon=1e-9):
    '''
    Args:
    B27 = 측점
    H20, H21, H22, H23 = 특정 스테이션 값
    D27 = 측거
    W13 = -곡선반경
    D13 = 캔트
    epsilon = tolerance for floating-point comparison

    Returns:
    A calculated value based on the conditions in the Excel formula.
    '''

    if H20 >= B27:
        return 0
    elif H21 > B27:
        return D27 - (D27 ** 5 / (40 * W13 ** 2 * D13 ** 2))
    elif H22 >= B27:
        return 0
    elif H23 >= B27:
        return D27 - (D27 ** 5 / (40 * W13 ** 2 * D13 ** 2))
    else:
        return 0

#이정량Y
def cal_spiral_shifty(B27, H20, H21, H22, H23, V27, W13, P11, epsilon=1e-9):
    '''
    Args:
    B27 = 측점
    H20, H21, H22, H23 = 특정 스테이션 값
    V27 = 이정량x
    W13 = -곡선반경
    P11 = x1
    epsilon = tolerance for floating-point comparison

    Returns:
    A calculated value based on the conditions in the Excel formula.
    '''

    if H20 >= B27:
        return 0
    elif H21 > B27:
        return V27 ** 3 / (6 * W13 * P11)
    elif H22 >= B27:
        return 0
    elif H23 >= B27:
        return V27 ** 3 / (6 * W13 * P11)
    else:
        return 0


def get_station_label(curve_type, station, SP_STA, PC_STA, CP_STA, PS_STA, EP_STA, epsilon=1e-9):
    """
    Args:
    curve_type (str): 곡선타입 ('simplecurve', 'cubic')
    station (float): 측점
    SP_STA, PC_STA, CP_STA, PS_STA, EP_STA (float): 각각 SP, PC, CP, PS, EP 좌표
    epsilon (float): 부동소수점 비교를 위한 허용 오차

    Returns:
    str: 해당 측점의 레이블 ('BC', 'SP', 'PC', 'CP', 'PS', 'EC', 'EP') or 빈 문자열 ("")
    """
    if abs(station - SP_STA) < epsilon:
        return 'BC' if curve_type == 'simplecurve' else 'SP'
    elif abs(station - PC_STA) < epsilon:
        return 'BC' if curve_type == 'simplecurve' else 'PC'
    elif abs(station - CP_STA) < epsilon:
        return 'EC' if curve_type == 'simplecurve' else 'CP'
    elif abs(station - PS_STA) < epsilon:
        return 'EC' if curve_type == 'simplecurve' else 'PS'
    elif abs(station - EP_STA) < epsilon:
        return "EP"
    else:
        return ""


#측거


def calculate_station_offset(B28, H20, H21, H22, H23, X13):
    '''
    Args:
    B28 = 측점 (Station point)
    H20, H21, H22, H23 = 스테이션 값 (Station values)
    X13 = BP STA (Base point)
    epsilon = tolerance for floating-point comparison

    Returns:
    A calculated offset value.
    '''
    #인트로 변환
    B28_int = int(B28)
    H20_int = int(H20)
    H21_int = int(H21)
    H22_int = int(H22)
    H23_int = int(H23)
    X13_int = int(X13)

    if B28_int == "":
        return ""
    elif B28_int <= H20_int:
        return B28 - X13
    elif B28_int < H21_int:
        return B28 - H20
    elif B28_int <= H22_int:
        return B28 - H21
    elif B28_int <= H23_int:
        return H23 - B28
    else:
        return B28 - H23


#선형 좌표계산함수
def calculate_coordinates(B27, H20, H21, H22, H23, V10, V11, W18, D27, W13, V27, W27, X14, X15, K20, N20, J9, N9, K23, N23):
    '''
    args
    B27 = 측점
    H20, H21, H22, H23 = 완화곡선 측점
    V10, V11 = 방위각 1,2
    W18 = sp, pc 방위각
    D27 = 측거
    W13 = -곡선반경
    
    V27 = 이정량 x
    W27 = 이정량 y
    X14, X15 = BP X, BP Y
    K20, K23 = SP[X], PS[X]
    J9, N9 = CENTER X, CENTER Y
    N20, N23 = SP[Y], PS[Y]
    '''

    #부동소수점 문제처리를 위해 int로 변경
    B27 = int(B27)#STA
    H20 = int(H20)#SP
    H21 = int(H21)#PC
    H22 = int(H22)#CP
    H23 = int(H23)#PS
    
    if B27 == "":
        return (0,0)

    # X coordinate calculation
    if B27 <= H20:#BP_SP
        X_result = math.cos(math.radians(V10)) * D27 + X14
    elif B27 < H21:#SP-PC
        X_result = (math.cos(math.radians(V10 + 90)) * W27 +
                    math.cos(math.radians(V10)) * V27 +
                    K20)
    elif B27 <= H22:#PC-CP
        X_result = (math.cos(math.radians(W18 - 90 + (180 * D27) / math.pi / W13)) * W13 +
                    J9)
    elif B27 <= H23:#CP-PS
        X_result = (math.cos(math.radians(V11 + 90)) * W27 +
                    math.cos(math.radians(V11 + 180)) * V27 +
                    K23)
    else:#PS-EP
        X_result = (math.cos(math.radians(V11)) * D27 + K23)

    # Y coordinate calculation
    if B27 <= H20:
        Y_result = math.sin(math.radians(V10)) * D27 + X15
    elif B27 < H21:
        Y_result = (math.sin(math.radians(V10 + 90)) * W27 +
                    math.sin(math.radians(V10)) * V27 +
                    N20)
    elif B27 <= H22:
        Y_result = (math.sin(math.radians(W18 - 90 + (180 * D27) / math.pi / W13)) * W13 +
                    N9)
    elif B27 <= H23:
        Y_result = (math.sin(math.radians(V11 + 90)) * W27 +
                    math.sin(math.radians(V11 + 180)) * V27 +
                    N23)
    else:
        Y_result = math.sin(math.radians(V11)) * D27 + N23

    return (X_result, Y_result)

def generate_integers(start, end, inc):
    # Round the start to the nearest increment
    start_rounded = (start + inc - 1) // inc * inc
    # Round the end to the nearest increment
    end_rounded = end // inc * inc
    
    return list(range(start_rounded, end_rounded + inc, inc))

    
def calculate_distance(x1, y1, x2, y2):
    # 거리계산함수
    distance = math.sqrt((x2 - x1)**2 + (y2 - y1)**2)
    distance_x = abs(x2 - x1)
    distance_y = abs(y2 - y1)
    return distance



def write_data(data):
    """
    Writes the given data to a CSV file in the specified format:
    type, sta, stationoffset, coordx, coordy, shiftx, shifty

    Args:
    - data (list of lists): The data to be written to the file. Each inner list represents a station's data.
    """
    file_path = 'C:/temp/data.csv'
    
    try:
        with open(file_path, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            
            # Write the header
            writer.writerow(['구분', '측점', '곡선거리', 'X', 'Y', '이정량x', '이정량y'])
            
            # Flatten the final_result and write data
            for station_group in data:
                for station_data in station_group:
                    for row in station_data:  # Iterate through the innermost list
                        if len(row) == 6:
                            lable, sta, station_offset, coord, shiftx, shifty = row
                            coordx, coordy = coord  # Unpack the coord tuple
                            
                            # Write the data into the CSV file
                            writer.writerow([lable, sta, station_offset, coordx, coordy, shiftx, shifty])
                        else:
                            print(f"Unexpected data format: {row}")
        
        print(f'Data successfully written to {file_path}')
    except Exception as e:
        print(f'An error occurred while writing to the file: {e}')

def write_report(data, data2):
    """
    Saves the alignment report to an Excel file, writing data into specific sheets using openpyxl.
    
    Parameters:
    data (list): The list of report variables for each curve.
    """

    #환경변수 설정
    km_precision = 2 #측점 자릿수
    coord_precision = 4 #좌표 자릿수
    dimention = 3#기타 숫자 자릿수
    filename = 'c:/temp/alignmentreport.xlsx'

    # Create a new workbook or load an existing one
    # Create a new workbook, overwriting any existing file
    workbook = Workbook()
    
    for i, entry in enumerate(data):
        # Create a new sheet or access the existing one dynamically based on i (IP1, IP2, ...)
        sheetname = f'IP{entry[0]}' # Dynamic sheet name like IP1, IP2, etc.

        # Check if the sheet already exists, if not create a new one
        if sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]
        else:
            sheet = workbook.create_sheet(title=sheetname)

        if entry[2] == 'simplecurve':  # Simple curve case
            
            sheet['A3'] = f"'===================〈 IP  NO. {entry[0]} >==================="#IPNO
            sheet['A4'] = f'I P [ X ] = {entry[1][0]:.{coord_precision}f}' # Write IP_X to cell F5
            sheet['D4'] = f'I P [ Y ] = {entry[1][1]:.{coord_precision}f}' # Write IP_Y to cell L5
            sheet['G6'] = '단곡선' #curvetype
            sheet['E6'] = f'({entry[3]})' #curvedirection
            sheet['A5'] = f"방위각 1 = {degrees_to_dms(entry[4])}" #방위각1
            sheet['E5'] = f"방위각 2 = {degrees_to_dms(entry[5])}" #방위각2
            sheet['A6'] = f'교각(IA) = {degrees_to_dms(entry[6])}'  #교각
            sheet['A7'] = 'CURVE CENTER'
            sheet['C7'] = f'X = {entry[7][0]:.{coord_precision}f}' # CENTER X
            sheet['F7'] = f'Y = {entry[7][1]:.{coord_precision}f}' # CENTER X
            sheet['A8'] = f'R = {entry[8]:.{dimention}f}' # R
            sheet['C8'] = f'IP 연장1 = {entry[9]:.{dimention}f}' #ip연장1
            
            sheet['A9'] = 'M = 0' #M
            sheet['C9'] = f'V = {entry[13]:.{dimention}f}' #V
            sheet['F9'] = f'Z = {entry[12] * 0.001:.{dimention}f}' #Z
            sheet['A10'] = f'TL = {entry[14][0]:.{dimention}f}' #TL
            sheet['C10'] = f'CL = {entry[14][1]:.{dimention}f}' #CL
            sheet['F10'] = f'SL = {entry[14][2]:.{dimention}f}' #SL
            sheet['A11'] = f'IP 정측점 = {format_distance(entry[18], decimal_places=km_precision)}' #IP정측점
            sheet['B12'] = f'BC = {format_distance(entry[19], decimal_places=km_precision)}' #BC측점
            sheet['B13'] = f'EC = {format_distance(entry[20], decimal_places=km_precision)}' #EC측점
            sheet['D12'] = f'X  = {entry[21][0]:.{coord_precision}f}' #BC_X
            sheet['F12'] = f'Y  = {entry[21][1]:.{coord_precision}f}' #BC_Y
            sheet['D13'] = f'X  = {entry[22][0]:.{coord_precision}f}' #EC_X
            sheet['F13'] = f'Y  = {entry[22][1]:.{coord_precision}f}' #EC_Y
            
        else:  # Cubic parabola case
            sheet['A3'] = f"'===================< IP  NO. {entry[0]} >==================="#IPNO
            sheet['A4'] = f'I P [ X ]  = {entry[1][0]:.{coord_precision}f}' # Write IP_X to cell F5
            sheet['D4'] = f'I P [ Y ]  = {entry[1][1]:.{coord_precision}f}' # Write IP_Y to cell L5
            sheet['G6'] = '3차포물선' #curvetype
            sheet['E6'] = f'({entry[3]})' #curvedirection
            sheet['A5'] = f"방위각 1  = {degrees_to_dms(entry[4])}" #방위각1
            sheet['E5'] = f"방위각 2  = {degrees_to_dms(entry[5])}" #방위각2
            sheet['A6'] = f'교각(IA)  = {degrees_to_dms(entry[6])}'  #교각
            sheet['A7'] = 'CURVE CENTER'
            sheet['C7'] = f'X  = {entry[7][0]:.{coord_precision}f}' # CENTER X
            sheet['F7'] = f'Y  = {entry[7][1]:.{coord_precision}f}' # CENTER X
            sheet['A8'] = f'R  = {entry[8]:.{dimention}f}' # R
            sheet['C8'] = f'IP 연장1  = {entry[9]:.{dimention}f}' #ip연장1
            
            sheet['A9'] = f'M  = {entry[11]:.{dimention}f}' #M
            sheet['C9'] = f'V  = {entry[13]:.{dimention}f}' #V
            sheet['F9'] = f'Z  = {entry[12] * 0.001:.{dimention}f}' #Z
            sheet['A12'] = f'X1  = {entry[14][0]:.{dimention}f}' #X1
            sheet['A10'] = f'TL  = {entry[14][10]:.{dimention}f}' #TL
            sheet['C10'] = f'CL  = {entry[14][12]:.{dimention}f}' #CL
            sheet['A11'] = f'LC  = {entry[14][11]:.{dimention}f}' #lc
            sheet['C12'] = f'Y1  = {entry[14][4]:.{dimention}f}' #Y1
            sheet['C11'] = f'L  = {entry[14][3]:.{dimention}f}' #L
            sheet['F10'] = f'SL  = {entry[14][13]:.{dimention}f}' #SL
            
            #sheet['H13'] = f'W = {entry[14][9]:.3f}' #W
            #sheet['L13'] = f'X2 = {entry[14][1]:.3f}' #X2
            #sheet['P13'] = f'S = {entry[14][7]:.3f}' #S
            #sheet['P14'] = f'F = {entry[14][6]:.3f}' #F
            sheet['D13'] = f'THITA = {degrees_to_dms(math.degrees(entry[17]))}' #PC점의 접선각
            #sheet['H16'] = f'Θ3 = {degrees_to_dms(entry[16])}' #SP>PC점의 방위각
            #sheet['M16'] = f'O→PC = {degrees_to_dms(entry[15])}' #O>PC점 방위각

            sheet['A13'] = f'B  = {degrees_to_dms(math.degrees(entry[14][16]))}'#B
            
            sheet['A14'] = f'C  = {degrees_to_dms(math.degrees(entry[14][15]))}'#C
            sheet['D14'] = f'원곡선교각  = {degrees_to_dms(math.degrees(entry[14][14]))}'#원곡선교각


            
            sheet['A15'] = f'IP 정측점  = {format_distance(entry[18], decimal_places=km_precision)}' #IP정측점
            sheet['D15'] = f'IP 역측점  = {format_distance(entry[18], decimal_places=km_precision)}' #IP정측점
            
            sheet['B16'] = f'SP  = {format_distance(entry[19], decimal_places=km_precision)}' #SP측점
            sheet['B17'] = f'PC  = {format_distance(entry[20], decimal_places=km_precision)}' #PC측점
            sheet['B18'] = f'CP  = {format_distance(entry[21], decimal_places=km_precision)}' #CP측점
            sheet['B19'] = f'PS  = {format_distance(entry[22], decimal_places=km_precision)}' #PS측점
            sheet['D16'] = f'X  = {entry[23][0]:.{coord_precision}f}' #SP_X
            sheet['F16'] = f'Y  = {entry[23][1]:.{coord_precision}f}' #SP_Y
            sheet['D17'] = f'X  = {entry[24][0]:.{coord_precision}f}' #PC_X
            sheet['F17'] = f'Y  = {entry[24][1]:.{coord_precision}f}' #PC_Y
            sheet['D18'] = f'X  = {entry[25][0]:.{coord_precision}f}' #CP_X
            sheet['F18'] = f'Y  = {entry[25][1]:.{coord_precision}f}' #CP_Y
            sheet['D19'] = f'X  = {entry[26][0]:.{coord_precision}f}' #PS_X
            sheet['F19'] = f'Y  = {entry[26][1]:.{coord_precision}f}' #PS_Y

        #공통
        sheet['B21'] = 'CHAINAGE'
        sheet['D21'] = 'X(North)'
        sheet['F21'] = 'Y(East)'

        #임시 방위각 출력
        '''
        sheet['G21'] = '접선방위각'
        sheet['H21'] = '사거리'
        sheet['I21'] = '편기각'
        sheet['J21'] = '접선각'
        sheet['K21'] = '방위각'
        '''
        
        # Write the relevant station data for the current entry
        if i < len(data2):  # Ensure you do not exceed data2 bounds
            station_group = data2[i]  # Get the corresponding station group for the current entry
            start_row = 22  # Start writing station data from row 21

            for station_data in station_group:
                for row in station_data:  # Iterate through the innermost list

                    lable, sta, station_offset, coord, shiftx, shifty , Ta, sageoriS , Q, T, Azimuth = row
                    # Write values into the sheet
                    sheet[f'A{start_row}'] = lable
                    sheet[f'B{start_row}'] = format_distance(sta, decimal_places=km_precision)
                    #sheet[f'C{start_row}'] = station_offset
                    sheet[f'D{start_row}'] = coord[0]
                    sheet[f'F{start_row}'] = coord[1]
                    
                    '''
                    sheet[f'G{start_row}'] = Ta #degrees_to_dms(Ta) #접선방위각
                    sheet[f'H{start_row}'] = sageoriS
                    sheet[f'I{start_row}'] = Q #degrees_to_dms(Q) #편기각
                    sheet[f'J{start_row}'] = T #degrees_to_dms(T) #접선각
                    sheet[f'K{start_row}'] = Azimuth #degrees_to_dms(Azimuth) #방위각
                    
                    #sheet[f'G{start_row}'] = shifty
                    '''
                    start_row += 1  # Move to the next row
            
    # Remove default sheet if it exists (usually "Sheet" when a new workbook is created)
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    # Save the workbook
    save_with_retry(workbook, filename, max_retries=100, delay=1)
    #workbook.save(filename)
    print(f"Data successfully written to {filename}")


def format_distance(number, decimal_places=2):
    negative = False
    if number < 0:
        negative = True
        number = abs(number)
        
    km = int(number) // 1000
    remainder = round(number % 1000, decimal_places)  # Round remainder to the specified decimal places
    
    # Format the remainder to have at least 'decimal_places' digits after the decimal point
    formatted_distance = "{:d}km{:0{}.{}f}".format(km, remainder, 4 + decimal_places, decimal_places)
    
    if negative:
        formatted_distance = "-" + formatted_distance
    
    return formatted_distance
'''
def format_distance(number, decimal_places=2):
    negative = False
    if number < 0:
        negative = True
        number = abs(number)
        
    km = int(number) // 1000
    remainder = "{:.2f}".format(number % 1000)
    formatted_distance = "{:d}km{:06.2f}".format(km, float(remainder))
    
    if negative:
        formatted_distance = "-" + formatted_distance
    
    return formatted_distance
'''

def get_station_coordinates(data1, data2, input_station):
    # input_station은 어느 IP에 속하는지 찾기

    for i, entry in enumerate(data1):
        IPN0 = entry[0]  # IP번호
        #print(f'data2길이 : {len(data2)}')
        station_group = data2[i]  # Get the corresponding station group for the current entry
        
        BP_STA = station_group[0][0][1]
        EP_STA = station_group[0][-1][1]

        BPXY = (station_group[0][0][3][0], station_group[0][0][3][1])

        
        '''
        print(f' BP_STA = {BP_STA}')
        print(f' EP_STA = {EP_STA}')
        print(f' BPXY = {BPXY}')
        '''
        
        
        
        
        try:
            h1 ,h2 = entry[4], entry[5]
            W13 = entry[8] if entry[3] == 'RIGHT CURVE' else entry[8] * -1

            
            if entry[2] == 'simplecurve':#단곡선
                BC_STA, EC_STA = entry[19], entry[20]
                
                if BP_STA<= input_station <= EP_STA:#해당IP인지 체크
                    
                    #맞으면 좌표계산 수행
                    
                    StationOffset = calculate_station_offset(input_station, BC_STA, BC_STA, EC_STA, EC_STA, BP_STA)#측거

                    BC_XY = entry[21]
                    EC_XY = entry[22]
                    
                    CURVE_CENTER = entry[7]
                    
                    coord = calculate_coordinates(input_station,
                                                  BC_STA, BC_STA, EC_STA, EC_STA,
                                                  h1, h2, h1,
                                                  StationOffset, W13,
                                                  0, 0,
                                                  BPXY[0], BPXY[1],
                                                  BC_XY[0], BC_XY[1],
                                                  CURVE_CENTER[0], CURVE_CENTER[1],
                                                  EC_XY[0], EC_XY[1])

                    if coord:#좌표를 찾은경우 반복문 탈출
                        break
                else:#해당 IP가 아니면 건너뛰고 다음 IP계산 수행
                    continue
                
            else:#3차포물선인경우
                
                SP_STA, PC_STA, CP_STA, PS_STA = entry[19], entry[20], entry[21], entry[22]
               
                StationOffset = calculate_station_offset(input_station, SP_STA, PC_STA, CP_STA, PS_STA, BP_STA)#측거
                
                if BP_STA<= input_station <= EP_STA:#해당IP인지 체크
                    
                    SP_PC_bearing = entry[16]
                    L = entry[14][3]
                    x1 = entry[14][0]
                    
                    #이정량X
                    shiftx =  cal_spiral_shiftx(input_station, SP_STA, PC_STA, CP_STA, PS_STA, StationOffset, W13, L)
                    #이정량Y
                    shifty =  cal_spiral_shifty(input_station, SP_STA, PC_STA, CP_STA, PS_STA, shiftx, W13, x1)

                    SP_XY = entry[23]
                    PS_XY = entry[26]
                    
                    CURVE_CENTER = entry[7]
                    
                    coord = calculate_coordinates(input_station,
                                              SP_STA, PC_STA, CP_STA, PS_STA,
                                              h1, h2, SP_PC_bearing,
                                              StationOffset, W13,
                                              shiftx, shifty,
                                              BPXY[0], BPXY[1],
                                              SP_XY[0], SP_XY[1],
                                              CURVE_CENTER[0], CURVE_CENTER[1],
                                              PS_XY[0], PS_XY[1])
                
                    if coord:#좌표를 찾은경우 반복문 탈출
                        break
                    
                else:#해당 IP가 아니면 건너뛰고 다음 IP계산 수행
                    continue
            if not coord:
                 print('측점이 범위를 벗어났습니다.')
                 return None
                
        except IndexError:
            print(f"Entry doesn't have enough data or is malformed: {entry}")
        except Exception as e:
            print(f"An error occurred: {e}")

    return coord

def create_civil3d_xlsx(data):
    filename = 'c:/temp/AlignmentExample.xlsx'

    # Create a new workbook or load an existing one
    # Create a new workbook, overwriting any existing file
    workbook = Workbook()
    sheetname = 'Sheet1' # Dynamic sheet name like IP1, IP2, etc.

    # Check if the sheet already exists, if not create a new one
    if sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
    else:
        sheet = workbook.create_sheet(title=sheetname)
            
    sheet['A1'] = 'Previous Entity ID'
    sheet['B1'] = 'Next Entity ID'
    sheet['C1'] = 'Radius'
    sheet['D1'] = 'Spiral Length'
    start_row = 2
    #data1에서 unwrap
    for i, entry in enumerate(data):
        R = entry[8]
        L = entry[14][3] if entry[2] == 'cubic' else ''
        sheet[f'A{start_row}'] = entry[0]
        sheet[f'B{start_row}'] = entry[0] +1
        sheet[f'C{start_row}'] = R
        sheet[f'D{start_row}'] = L

        start_row += 1  # Move to the next row
        
    # Remove default sheet if it exists (usually "Sheet" when a new workbook is created)
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    # Save the workbook
    save_with_retry(workbook, filename, max_retries=100, delay=1)
    #workbook.save(filename)
    print(f"Data successfully written to {filename}")



def find_perpendicular_projection(coord, linestring):

    # coord가 (x, y) 튜플일 경우 Point 객체로 변환
    if not isinstance(coord, Point):
        point = Point(coord)  # 튜플을 Point 객체로 변환
    else:
        point = coord
    
    # linestring이 Shapely LineString 객체인지 확인
    if not isinstance(linestring, LineString):
        print("linestring은 Shapely LineString 객체여야 합니다.")
        #Shapely LineString 객체로 변환
        linestring = LineString(linestring)
        
    # Project the point onto the LineString
    projected_distance = linestring.project(point)
    
    
    # 점을 linestring에 투영
    projection_point = linestring.interpolate(projected_distance)
    
    # 투영된 점의 거리에 해당하는 선형 길이 찾기
    projected_length = linestring.project(point)

    # 투영된 점이 속한 segment의 시작 좌표를 찾기 위해 각 세그먼트를 순회
    for i, (start, end) in enumerate(zip(linestring.coords[:-1], linestring.coords[1:])):
        segment = LineString([start, end])
        segment_length = segment.length
        
        # 투영된 거리가 현재 segment 내에 있는지 확인
        if projected_length <= segment_length:
            start_coordinate = Point(start)
            break
        else:
            projected_length -= segment_length
    
    # 점과 투영된 좌표 간의 거리 (오프셋 거리)
    offset_distance = point.distance(projection_point)

    # 투영된 좌표와 시작 좌표 간의 수평 거리
    horizen_distance = projection_point.distance(start_coordinate)
        
    # 결과 반환: [오프셋 거리, 투영된 좌표, 수평 거리, 시작 좌표]
    result = [offset_distance, projection_point, horizen_distance, start_coordinate]

    print(f' projection_point = {projection_point}')
    
    return result

def plot_projection(point, projection_point, linestring):

    # linestring이 Shapely LineString 객체인지 확인
    if not isinstance(linestring, LineString):
        print("linestring은 Shapely LineString 객체여야 합니다.")
        #Shapely LineString 객체로 변환
        linestring = LineString(linestring)
        
    # coord가 (x, y) 튜플일 경우 Point 객체로 변환
    if not isinstance(point, Point):
        point = Point(point)  # 튜플을 Point 객체로 변환
    else:
        point = coord    
    # Extract x and y coordinates from LineString
    x, y = linestring.xy

    # Create a plot
    plt.figure(figsize=(6, 6))
    
    # Plot LineString
    plt.plot(x, y, label="LineString", color='blue', linewidth=2)

    # Plot original point
    plt.plot(point.x, point.y, 'ro', label="Original Point", markersize=8)

    # Plot projection point
    plt.plot(projection_point.x, projection_point.y, 'go', label="Projection Point", markersize=8)

    # Draw a line between the original point and the projection point
    plt.plot([point.x, projection_point.x], [point.y, projection_point.y], 'k--', label="Perpendicular Line")

    # Add labels
    plt.xlabel("X")
    plt.ylabel("Y")
    plt.axis('equal')
    
    plt.legend()

    # Show the plot
    plt.grid(True)
    plt.show()
    
def get_station_from_coordinates(linestring, coord, data):
    tolerance = 1e-6
    
    #찾을 station
    find_station = None

    #토목좌표를 수학좌표로 변환
    # 각 튜플의 요소들을 뒤집기(토목좌표를 수학좌표로 변환)
    linestring_math = [(y, x) for x, y in linestring]

    
    #튜플뒤집기(토목좌표를 수학좌표로 변환)
    coord_math = (coord[1],coord[0])

    
    
    #주어진 점을 선분에 수직으로 투영하는 함수호출
    result2 = find_perpendicular_projection(coord_math, linestring_math)#입력좌표,polycurve

    distance = result2[0]#offset 거리
    horizen_distance = result2[2]#시작점에서 투영점까지의 거리
    
    # result2[3]이 Point 객체라면 x, y 속성을 사용하여 튜플로 변환
    start_coordinate = (result2[3].x, result2[3].y)  # Point 객체에서 x, y 좌표 추출
    perpendicular_coord = (result2[1].x, result2[1].y)  # Point 객체에서 x, y 좌표 추출

    
    #선분 시작점의 좌표를 data 리스트에서 찾기
    for i , group in enumerate(data):
        for j, entry in enumerate(group):
            for k, row in enumerate(entry):
                destinate_coord = row[3][1]
                
                if abs(start_coordinate[0] - destinate_coord) < tolerance:  # e.g., tolerance = 1e-6
                    find_station = row[1]
                    print(f"선형계산서 측점: {find_station}")
                    break
    
    if find_station:
        final_station = horizen_distance + find_station
        print(f"최종 측점: {final_station}")
        print(f'offset: {distance:.2f}')
        print(f'좌표: X = {perpendicular_coord[0]:.4f}, Y = {perpendicular_coord[1]:.4f}')     
        
    else:
        print("sta 값을 찾을 수 없습니다.")
        final_station = 0


    plot_projection(coord_math, result2[1], linestring_math)
    return final_station

def main():
    global unit
    #변수입력받기
    user_input1 = input('계산 간격 입력: (기본값 20) ')
    unit = int(user_input1) if user_input1 else 20
    
    user_input2 = input('시작 측점 입력: (기본값 0) ')
    start_STA = float(user_input2) if user_input2 else 0
    
    user_input3 = input('설계최고속도 입력: (기본값 250) ')
    designSpeed = int(user_input3) if user_input3 else 250
    
    user_input4 = input('구간내 적용 도상입력 0자갈도상 1콘크리트도상 : (기본값 자갈도상) ')
    
    if user_input4 ==0:
        ballast = '자갈도상'
    else:
        ballast = '콘크리트도상'
    
    user_input5 = input('M,Z,V 수동입력?: y or n (기본값 아니오) Z단위:MM ').strip().lower()
    
    # Check if the input is "yes" or "y" (indicating manual input)
    ismanualmzv = user_input5 in ['y', 'yes']
    
    lines = read_file()
    if not lines:
        print('입력된 선형 없음')
        return
    
    alignment = pasing_ip(lines)
    linestring = create_Linestring(alignment)
    
    Radius_list = [point[2] for point in alignment[1:-1]]
    
    ia_list = calculate_angles_and_plot(linestring)

    if user_input5:#수동입력
        parameters = []
        for i in range(len(Radius_list)):
            inputvariable = input('Enter M,Z,V (쉼표로 구분): ').split(',')  # Split the input
            print(f'현재 입력값 IP{i+1}  M = {inputvariable[0]}, Z = {inputvariable[1]} , V= {inputvariable[2]}')
            
            
            inputvariable = [float(x) for x in inputvariable]  # Convert each value to float
            parameters.append(inputvariable)
        #잠시 대기
        print('입력이 완료되었습니다. 계산을 수행합니다.')    
    else:#자동입력
        parameters = cal_parameter(Radius_list, designSpeed, ballast)
    
    
    final_result, alignment_report_variable_list , polycurve = calculate_curve(linestring, Radius_list, ia_list, parameters, unit, start_STA)
    
    #좌표출력
    #write_data(final_result)
    #print(final_result)
    
    #선형계산서 출력
    write_report(alignment_report_variable_list, final_result)
    print('선형계산서 출력완료')

    #CIVIL3D성과물 생성
    create_civil3d_xlsx(alignment_report_variable_list)
    print('CIVIL3D 성과물 출력완료')
    
    a= 0

    
    while 1:
        
        
        print('원하는 작업을 선택하세요')
        print('1. 도면출력')
        print('2. 측점으로 좌표계산')
        print('3. 좌표로 측점찾기')
        print('4. 프로그램 종료')
        
        a =  int(input(' 번호 입력: '))
        
        if a == 1:
            print('선택한 메뉴 : 도면출력')
            while True:
                user_input10 = input('도면축척 입력 (기본값 1:1000): ')  # Get input as a string
                if user_input10.isdigit() or user_input10 == '':
                    scale = int(user_input10) if user_input10 else 1000
                    print(f'현재 도면축척 1:{scale}')                 
                    create_dxf(polycurve, linestring, alignment_report_variable_list, final_result, filename='본선평면선형.dxf', scale=scale)
                    print('도면출력이 완료됐습니다.')
                    break  # Exit the loop after successful creation
                else:
                    print("유효하지 않은 입력입니다. 숫자를 입력해주세요.")
                
        elif a ==2:
            print('선택한 메뉴 : 측점으로 좌표계산')
            
            while 1:
                input_station = float(input('측점 입력: '))

                
                find_coord = get_station_coordinates(alignment_report_variable_list, final_result, input_station)

                if find_coord:
                    print(f'찾은 좌표 X = {find_coord[0]:.4f}, Y = {find_coord[1]:.4f}')

                

                    print('계산이 종료되었습니다. 다른 측점 계산은 1, 이전 메뉴로 돌아가려면 2를 입력하세요')
                    number = int(input('번호 입력: '))
                    if number == 1:
                        continue
                    elif number ==2:
                        break
                else:#좌표를 못찾은경우 다른 측점 입력해야함.
                    print('좌표를 못찾았습니다. 다른 측점을 입력하세요')
                    continue
        elif a == 3:
            print('선택한 메뉴: 좌표로 측점찾기')
            
            while 1:
                # 사용자로부터 좌표 입력을 받음
                input_coordinates = input('좌표 입력 (쉼표로 구분): ')
                
                # 입력받은 좌표를 쉼표로 구분하여 X, Y 값을 실수로 변환
                try:
                    x, y = map(float, input_coordinates.split(','))
                    
                    # 좌표로 측점 찾는 로직을 여기에 추가
                    print(f'입력된 좌표: X = {x}, Y = {y}')
                    #좌표를 튜플로 묶음
                    input_xy = (x,y)
                    
                    # 이후 처리 로직 (예: 좌표를 사용하여 측점 계산 함수 호출)
                    print('통과')
                    # 중복 제거
                    
                    unique_data = list(set(polycurve))
                    
                    find_station = get_station_from_coordinates(unique_data, input_xy, final_result)
                    print(f'찾은 측점: {format_distance(find_station)}')
                    
                    print('계산이 종료되었습니다. 다른 좌표 계산은 1, 이전 메뉴로 돌아가려면 2를 입력하세요')
                    number = int(input('번호 입력: '))
                    if number == 1:
                        continue
                    elif number ==2:
                        break

                except ValueError as e:
                    print(e)
             
        elif a== 4:
            print('프로그램 종료')
            break
if __name__ == '__main__':
    main()
