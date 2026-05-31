import random
import copy
import math
import csv
import matplotlib.pyplot as plt
import time
from tkinter import filedialog
import openpyxl

#전역변수
chain =25
isstartend = True # True 이면 존재 False 면 미존재
destination_score = 100 #목표점수
similarity_threshold = 5 #예측 정확도

start_break_point = 1000 #시작 vip고정
end_break_point = 11550 #종점 vip고정

# 생성할 대안 갯수
alternative_count = 'inf'

if isstartend:
    detail_design = [(0,51),(12350,46)]
else:
    detail_design = [(0,0)]

plt.rcParams['font.family'] ='Malgun Gothic'
plt.rcParams['axes.unicode_minus'] =False

# csv 파일에서 지반고 정보와 측점 정보를 불러와서 gl 배열에 저장
def load_gl():
    gl = []
    file_path = filedialog.askopenfilename(title="파일 선택", filetypes=[("텍스트 파일", "*.txt")])

    if file_path:
        try:
            with open(file_path, 'r',encoding='UTF-8') as file:
                reader = csv.reader(file)
                next(reader)  # 첫 번째 행 건너뛰기
                for row in reader:
                    station = int(row[0]) # 측점 정보
                    ground_elevation = float(row[1])  # 지반고 정보
                    
                    gl.append([station, ground_elevation])
                    
        except FileNotFoundError:
            print("파일을 찾을 수 없습니다.")
        except Exception as e:
            print("파일을 읽는 중 오류가 발생했습니다:", e)
    else:
        print("파일 선택이 취소되었습니다.")


    return gl

def detect_terrain(gl):
    # 경사도를 저장할 리스트
    slope_type = []

    # 고도 데이터를 반복하여 경사도를 계산
    for i in range(1, len(gl) - 1):
        prev_elevation = gl[i - 1][1]
        current_elevation = gl[i][1]
        next_elevation = gl[i + 1][1]

        # 현재 고도와 이전, 다음 고도의 차이를 계산
        diff_prev = current_elevation - prev_elevation
        diff_next = next_elevation - current_elevation

        # 거리 변화가 0이 아닌 경우에만 경사도를 계산하여 추가
        if diff_prev != 0 and diff_next != 0:
            slope_prev = (diff_prev / chain) * 100  # 이전 측점에서 현재 측점까지의 경사도
            slope_next = (diff_next / chain) * 100  # 현재 측점에서 다음 측점까지의 경사도
            average_slope = (slope_prev + slope_next) / 2  # 두 경사도의 평균을 사용하여 현재 측점의 경사도 계산

            # 평균 경사도에 따라 경사도 유형을 할당
            if average_slope < 3:
                slope_type.append('평지')
            elif 3 <= average_slope < 10:
                slope_type.append('완만한 경사')
            elif 10 <= average_slope < 20:
                slope_type.append('중간 경사')
            else:
                slope_type.append('급한 경사')
        else:
            slope_type.append('N/A')  # 경사도를 계산할 수 없는 경우

    return slope_type


def generate_random_profile(num_points, min_distance, gl):
    # 초기 설정
    start_station, start_elevation = gl[0]
    end_station, end_elevation = gl[-1]

    detail_start_station, detail_start_elevation = detail_design[0]
    detail_end_station, detail_end_elevation = detail_design[-1]
    
    # 시작점 설정
    if not isstartend: #False
        points = [[start_station, start_elevation + 10]]  # 초기 계획고 설정
    else:
        points = [[detail_start_station, detail_start_elevation]]

    #다음점은 LEVEL로 맞춰야함
    
    # 현재 위치 설정
    current_station = start_station
    current_elevation = start_elevation + 10

    for i in range(num_points - 1):
        # 다음 지점까지의 거리 설정
        distance_to_next = chain * math.ceil(random.uniform(min_distance,min_distance * 2) / chain)

        if current_station + distance_to_next >= end_station:
            break

        # 다음 지점 설정
        next_station = current_station + distance_to_next
        
        # 다음 지점의 지반고 찾기
        next_elevation = None

        if isstartend:
            for station, elevation in gl:
                if station == next_station:
                    next_elevation = elevation + 10
                    break
            if i == 0:
                next_elevation = detail_start_elevation
        else:
            for station, elevation in gl:
                if station == next_station:
                    next_elevation = elevation + 10
                    break


        # 다음 지점의 지반고가 없는 경우 예외 처리
        if next_elevation is None:
            raise ValueError(f"No ground elevation data found for station {next_station}")

        # 현재 위치 업데이트
        current_station = next_station
        current_elevation = next_elevation

        # 지점 추가
        points.append([current_station, current_elevation])

    # 마지막 지점 설정
    
    if isstartend:
        points.append([end_station, detail_end_elevation])
        points[-2][0] = end_break_point
        points[-2][1] = points[-1][1]
    else:
        points.append([end_station, end_elevation + 10])
    return points


def check_and_adjust_elevation(profile):
    adjusted_profile = []
    for i, (station, elevation) in enumerate(profile):
        rand_el = random.uniform(0, 20)
        if isstartend:
            if i == 1:
                station = start_break_point
            elif i > 1 and i < len(profile) - 3:
                prev_station, prev_elevation = adjusted_profile[-1]
                if abs(elevation - prev_elevation) > 20:
                    elevation = prev_elevation + (rand_el if elevation > prev_elevation else -rand_el)

        else:
            if i > 0:
                prev_station, prev_elevation = adjusted_profile[-1]
                if abs(elevation - prev_elevation) > 20:
                    elevation = prev_elevation + (rand_el if elevation > prev_elevation else -rand_el)
        adjusted_profile.append([station, elevation])

    return adjusted_profile

'''
def check_and_adjust_elevation(points, slope_type, gl):
    adjusted_profile = []

    for i, (station, elevation) in enumerate(points):
        gl_station, gl_elevation = gl[i]
        if slope_type == "평지":  # 평지인 경우
            rand_el = random.uniform(-10, 10)
            adjusted_elevation = gl_elevation + rand_el
        else:  # 평지가 아닌 경우
            if i == 0:  # 첫 번째 측점인 경우
                adjusted_elevation = elevation
            else:
                prev_station, prev_elevation = points[i - 1]
                if abs(elevation - prev_elevation) > 20:  # 이전 측점과의 높이 차이가 20 이상인 경우
                    rand_el = random.uniform(0, 10)
                    adjusted_elevation = prev_elevation + (rand_el if elevation > prev_elevation else -rand_el)
                else:  # 이전 측점과의 높이 차이가 20 이하인 경우
                    rand_el = random.uniform(-10, 10)
                    adjusted_elevation = elevation + rand_el
        adjusted_profile.append([station, adjusted_elevation])

    return adjusted_profile
'''

#최급기울기 찾기
def find_steepest_gradient(g):
    max_gradient = 0
    steepest_gradient_station = None
    
    for i in range(len(g) - 1):
        station1, elevation1 = g[i]
        station2, elevation2 = g[i + 1]
        
        # 두 점 사이의 거리
        distance = station2 - station1
        
        if distance == 0:
            continue
        
        # 기울기 계산
        gradient = abs((elevation2 - elevation1) / distance) * 1000
        
        # 최대 기울기 갱신
        if gradient > max_gradient:
            max_gradient = gradient
            steepest_gradient_station = station1 if elevation1 < elevation2 else station2
            
    return max_gradient, len(g) - 2

#표고점을 20간격으로 계획고 생성
def generate_station_elv(points):
    station_elv = []

    for i in range(len(points) - 1):
        start_station = points[i][0]
        start_elevation = points[i][1]
        end_station = points[i + 1][0]
        end_elevation = points[i + 1][1]

        
        height = end_elevation - start_elevation
        distance = end_station - start_station

    
        if distance == 0:
            slope = 0
        else:
            slope = height / distance * 1000

        num_steps = int(distance / chain)  # 20의 배수인 측점 개수 계산

        for j in range(num_steps + 1):
            current_station = start_station + j * chain
            current_elevation = start_elevation + slope/1000 * (current_station - start_station)
            station_elv.append([current_station, current_elevation])

    return station_elv

#절토성토고 계산
def calculate_cutfill(station_elv, gl):
    new_cutfill = []
    
    for gl_station, el in gl:  # 지반고 추출
        for elv_station, dl in station_elv:  # 계획고 추출
            if gl_station == elv_station:
                height_d = el - dl
                new_cutfill.append([gl_station, height_d])

    return new_cutfill

#구조물판별
def cal_profile_structure(new_cutfill):
    new_structure = []

    for station, cutfill_d in new_cutfill:
        if cutfill_d  >= 12:
             new_structure.append([station , "터널"])
            
        elif cutfill_d  <= -12:
            new_structure.append([station , "교량"])
        else:
             new_structure.append([station , "토공"])

    return new_structure

#구조물 측점
def find_structure_ranges(structure_list):
    ranges = {}
    bridge_count = 1
    tunnel_count = 1

    current_structure = None
    start_station = None

    for i, (station, structure) in enumerate(structure_list):
        if structure == '교량':
            if current_structure != '교량':
                if current_structure is not None:
                    if current_structure == '터널':
                        key = f'T{tunnel_count}'
                        ranges[key] = (start_station, structure_list[i - 1][0])
                        tunnel_count += 1
                start_station = station
                current_structure = '교량'
        elif structure == '터널':
            if current_structure != '터널':
                if current_structure is not None:
                    if current_structure == '교량':
                        key = f'B{bridge_count}'
                        ranges[key] = (start_station, structure_list[i - 1][0])
                        bridge_count += 1
                start_station = station
                current_structure = '터널'
        else:
            if current_structure == '교량':
                key = f'B{bridge_count}'
                ranges[key] = (start_station, structure_list[i - 1][0])
                bridge_count += 1
            elif current_structure == '터널':
                key = f'T{tunnel_count}'
                ranges[key] = (start_station, structure_list[i - 1][0])
                tunnel_count += 1
            current_structure = None
            start_station = None

    if current_structure == '교량':
        key = f'B{bridge_count}'
        ranges[key] = (start_station, structure_list[-1][0])
    elif current_structure == '터널':
        key = f'T{tunnel_count}'
        ranges[key] = (start_station, structure_list[-1][0])

    return ranges

# 구조물 갯수
def count_structures(structure_ranges):
    count_bridge = 0
    count_tunnel = 0

    for key in structure_ranges.keys():
        if key.startswith('B'):
            count_bridge += 1
        elif key.startswith('T'):
            count_tunnel += 1

    return count_bridge, count_tunnel

# 공사비용 계산 함수
def cal_profile_money(new_structure):
    length_bridge = 0
    length_tunnel = 0
    length_excavation = 0

    for station, structure_type in new_structure:
        if "교량" in structure_type:
            length_bridge += 20
        if "터널" in structure_type:
            length_tunnel += 20
        if "토공" in structure_type:
            length_excavation += 20

    cost_bridge = (length_bridge / 1000) * 24574
    cost_tunnel = (length_tunnel / 1000) * 15784
    cost_excavation = (length_excavation / 1000) * 8620

    new_total_cost = cost_bridge + cost_tunnel + cost_excavation

    return new_total_cost

#노선 점수계산
def fitness(points,elv):
    score = 0  # Initial score is 0
    max_score = 100  # Maximum score, arbitrary value
    
    for i in range(len(points) - 1):
        start_station = points[i][0]
        start_elevation = points[i][1]
        end_station = points[i + 1][0]
        end_elevation = points[i + 1][1]

        # 변수설정
        height = end_elevation - start_elevation
        distance = end_station - start_station

        if distance == 0:
            slope = 0
        else:
            slope = height / distance * 1000

        #최소구배길이 미달시 0점
        if distance <= 200:
            return score
        
    
        len_score = 0.1
        score += len_score

        #구배점수 계산
        slope_score = 0.1
        if slope <= 25:
            score += slope_score
        else:
            score = 0
        #절성토고 점수
        cutfill_score = 0.1
        
        for _, elvation in elv: 
            if elvation <= 25:
                score += cutfill_score
            else:
                score -= cutfill_score
    # Return the score
    return (score / max_score) * 100

def create_excel(final_alternative):
    # 엑셀 파일 생성
    wb = openpyxl.Workbook()

    # '종단' 시트 생성
    ws1 = wb.active
    ws1.title = "종단"

    # 데이터 쓰기
    # 헤더 추가
    ws1.append(["측점", "계획고"])
    for station, elevation in final_alternative['종단']:
        ws1.append([station, elevation])

    # '구조물' 시트 생성
    ws2 = wb.create_sheet(title="교량")
    ws3 = wb.create_sheet(title="터널")
    # 데이터 쓰기
    # 헤더 추가
    ws2.append(["명칭", "시점", "종점","길이"])
    ws3.append(["명칭", "시점", "종점","길이"])
    
    for struct_type, (start_station, end_station) in final_alternative['구조물위치'].items():
        length = end_station - start_station
        if struct_type.startswith('B'):
            ws2.append([struct_type, start_station, end_station,length])
        else:
            ws3.append([struct_type, start_station, end_station,length])
    
    # '절성고' 시트 생성
    ws4 = wb.create_sheet(title="절성고")

    # 데이터 쓰기
    # 헤더 추가
    ws4.append(["측점", "절성토"])
    
    for station,diff in final_alternative['절성고']:
        ws4.append([station, diff*-1])
    
    # 엑셀 파일 저장
    wb.save('c:\\temp\\최적대안.xlsx')


    
#테스트코드
gl = load_gl()

#vip간 최소거리
min_distance = 1000

results = []



#최대 vip갯수
max_vip = int(gl[-1][0] / min_distance)
print(max_vip)



#무한루프를 돌면서 각 대안별 항목 계산

#목표점수 변수
is_destination_score = False


i=0
while True:
    if i > 1 and i % 200 == 0:
        destination_score -= 5
        print(f"목표점수 {destination_score:.0f}점")
    print(f'\n------ 현재 대안 {i+1} -----\n')
    num_points = max_vip
    alternative = {}
    
    terrain = detect_terrain(gl)
    
    alternative['종단'] = generate_random_profile(num_points, min_distance, gl)
    #alternative['종단'] = check_and_adjust_elevation(alternative['종단'], terrain, gl)
    alternative['종단'] = check_and_adjust_elevation(alternative['종단'])
    
    alternative['최급기울기'] = find_steepest_gradient(alternative['종단'])
    alternative['계획고'] = generate_station_elv(alternative['종단'])
    alternative['절성고'] = calculate_cutfill(alternative['계획고'], gl)
    alternative['구조물'] = cal_profile_structure(alternative['절성고'])
    alternative['구조물위치'] = find_structure_ranges(alternative['구조물'])
    alternative['구조물갯수'] = count_structures(alternative['구조물위치'])
    alternative['공사비'] = cal_profile_money(alternative['구조물'])
    alternative['점수'] = fitness(alternative['종단'],alternative['절성고'])

    print(f'점수= {alternative["점수"]}')
    print(f'공사비 = {alternative["공사비"]:.2f} 백만원')
    print(f'최급기울기 = {alternative["최급기울기"][0]:.2f} 퍼밀')
    print(f'기울기개소 = {alternative["최급기울기"][1]} 개소')
    print(f'구조물 = (교량: {alternative["구조물갯수"][0]}개소, 터널: {alternative["구조물갯수"][1]}개소)')

    results.append(alternative)

    for j, alt in enumerate(results):
        alt_score = alt['점수']
        alt_cost= alt['공사비']
        alt_maxgrade = alt['최급기울기'][0]
        alt_numgrade = alt['최급기울기'][1]
        alt_struct = alt['구조물위치']
        alt_numbr = alt['구조물갯수'][0]
        alt_numtn = alt['구조물갯수'][1]
        alt_fl_list = alt['계획고']
        difference = abs(alt_score - destination_score)
        
        if difference <= similarity_threshold and alt_maxgrade < 25:
            print(f'\n-----최적 대안 {j+1}-----\n')
            print(f"목표점수 {destination_score:.0f}점")
            print(f'최적대안 점수= {alt_score:.0f}점')
            print(f'최적대안 공사비 : {alt_cost:.2f} 백만원')
            print(f'최적대안 최급기울기 = {alt_maxgrade:.2f} 퍼밀')
            print(f'최적대안 기울기개소 = {alt_numgrade} 개소')
            print(f'최적대안 구조물 = (교량: {alt_numbr}개소, 터널: {alt_numtn}개소)')

            is_destination_score = True
            final_alternative_index = j
            break

    if is_destination_score:
        break

    i += 1
if is_destination_score:
    # 최적 대안을 그래프로 그리기
    final_alternative = results[final_alternative_index]
    stations, elevations = zip(*final_alternative['종단'])

    #기존 gl
    glx, gly = zip(*gl)
    plt.plot(glx,gly, linestyle='-', color='k',label='GL')

    #최적대안
    plt.plot(stations, elevations, marker='o', linestyle='-', color='red', label='final Line')

    #기울기 표시
    # 각 계획선의 중심에 기울기 표시
    i= 0
    for i in range(len(stations) - 1):
        x1, y1 = stations[i], elevations[i]
        x2, y2 = stations[i + 1], elevations[i + 1]
        
        
        gradient = (y2 - y1) / (x2 - x1) * 1000 # 기울기 계산
        if gradient !=0:
            gradient_text = f'{gradient:.2f}'.rstrip('0').rstrip('.')
        else:
            gradient_text = 'L'  # 분모가 0인 경우 'L'로 설정
        
        midpoint = (x1 + x2) / 2, (y1 + y2) / 2   + 10# 두 점 사이의 중심

        plt.text(midpoint[0], midpoint[1], gradient_text, fontsize=10, color='red', ha='center', va='center')

    #구조물 시종점 표시
    # 교량 시종점에 수직선 추가
    for bridge_key, (start_station, end_station) in alt_struct.items():
        for station, elevation in alt_fl_list:
            if bridge_key.startswith('B') and start_station == station:
                plt.plot([start_station, start_station], [elevation, elevation + 100], color='skyblue', linestyle='-')
                plt.text(start_station, elevation + 100, bridge_key, color='r')
            elif bridge_key.startswith('B') and end_station == station:
                plt.plot([end_station, end_station], [elevation, elevation + 100], color='skyblue', linestyle='-')

            elif bridge_key.startswith('T') and start_station == station:
                plt.plot([start_station, start_station], [elevation, elevation + 100], color='pink', linestyle='-')
                plt.text(start_station, elevation + 100, bridge_key, color='r')
            elif bridge_key.startswith('T') and end_station == station:
                plt.plot([end_station, end_station], [elevation, elevation + 100], color='pink', linestyle='-')
            else:
                continue
    # 결과를 엑셀파일로 저장
    create_excel(final_alternative)

    # Set labels and title
    plt.xlabel('Station')
    plt.ylabel('FL (Finish Line) Elevation')
    plt.title('Alternative Alignment')
    plt.legend()
    plt.axis('auto')
    # Show the plot
    plt.show()

    
                
# 코드 실행 끝
print('EOF')
