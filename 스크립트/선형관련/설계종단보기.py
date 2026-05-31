import random
import copy
import math
import csv
import matplotlib.pyplot as plt
from tkinter import filedialog, Tk, StringVar, ttk
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import numpy as np
import openpyxl
import tkinter as tk
from matplotlib.figure import Figure

# 전역변수
chain = 40

plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

def gui_thread():
    root = tk.Tk()
    label = tk.Label(root, text="GUI 메인 루프를 실행합니다.")
    label.pack()
    root.mainloop()

# csv 파일에서 지반고 정보와 측점 정보를 불러와서 gl 배열에 저장
def load_gl():
    gl = []
    file_path = filedialog.askopenfilename(title="파일 선택", filetypes=[("텍스트 파일", "*.txt")])

    if file_path:
        try:
            with open(file_path, 'r', encoding='UTF-8') as file:
                reader = csv.reader(file)
                #next(reader)  # 첫 번째 행 건너뛰기
                for row in reader:
                    station = int(row[0])  # 측점 정보
                    ground_elevation = float(row[1])  # 지반고 정보
                    
                    gl.append([station, ground_elevation])
                    
        except FileNotFoundError:
            print("파일을 찾을 수 없습니다.")
        except Exception as e:
            print("파일을 읽는 중 오류가 발생했습니다:", e)
    else:
        print("파일 선택이 취소되었습니다.")

    return gl

# csv 파일에서 계획고 정보와 측점 정보를 불러와서 fl 배열에 저장
def load_fl():
    fl = []
    file_path = filedialog.askopenfilename(title="파일 선택", filetypes=[("텍스트 파일", "*.txt")])

    if file_path:
        try:
            with open(file_path, 'r', encoding='UTF-8') as file:
                reader = csv.reader(file)
                  # 첫 번째 행 건너뛰기
                for row in reader:
                    station = int(row[0])  # 측점 정보
                    design_elevation = float(row[1])  # 계획고 정보
                    
                    fl.append([station, design_elevation])
                    
        except FileNotFoundError:
            print("파일을 찾을 수 없습니다.")
        except Exception as e:
            print("파일을 읽는 중 오류가 발생했습니다:", e)
    else:
        print("파일 선택이 취소되었습니다.")

    return fl

# 최급기울기 찾기
def find_steepest_gradient(g):
    max_gradient = 0
    steepest_gradient_station = None
    
    for i in range(len(g) - 1):
        station1, elevation1 = g[i]
        station2, elevation2 = g[i + 1]
        
        # 두 점 사이의 거리
        distance = station2 - station1
        
        if distance == 0:
            continue
        
        # 기울기 계산
        gradient = abs((elevation2 - elevation1) / distance) * 1000
        
        # 최대 기울기 갱신
        if gradient > max_gradient:
            max_gradient = gradient
            steepest_gradient_station = station1 if elevation1 < elevation2 else station2
            
    return max_gradient, len(g) - 2

def generate_station_elv(points):
    station_elv = []

    for i in range(len(points) - 1):
        start_station = points[i][0]
        start_elevation = points[i][1]
        end_station = points[i + 1][0]
        end_elevation = points[i + 1][1]

        height = end_elevation - start_elevation
        distance = end_station - start_station

        if distance == 0:
            slope = 0
        else:
            slope = height / distance * 1000

        num_steps = distance // chain  # 40의 배수인 측점 개수 계산

        for j in range(num_steps):
            current_station = start_station + j * chain
            current_elevation = start_elevation + slope / 1000 * (current_station - start_station)
            if len(station_elv) == 0 or station_elv[-1][0] != current_station:
                station_elv.append([current_station, current_elevation])

        # 마지막 지점 추가
        if len(station_elv) == 0 or station_elv[-1][0] != end_station:
            station_elv.append([end_station, end_elevation])

    return station_elv

# 절토성토고 계산
def calculate_cutfill(station_elv, gl):
    new_cutfill = []
    
    for (gl_sta, gl) , (fl_sta,fl) in zip(gl ,station_elv):  # 지반고 추출
        height_d = gl - fl
        new_cutfill.append([gl_sta, height_d])

    return new_cutfill

# 구조물 판단
def structure_judgement(new_cutfill):
    new_structure = []

    for station, cutfill_d in new_cutfill:
        if cutfill_d >= 15:
            structure_type = '교량'
        elif cutfill_d <= -15:
            structure_type = '터널'
        else:
            structure_type = 'N/A'

        new_structure.append([station, structure_type])

    return new_structure

# 토량을 구하고 점수 매기기
def calculate_cutfill_volume_and_score(new_cutfill):
    total_cut_volume = 0
    total_fill_volume = 0

    for station, cutfill_d in new_cutfill:
        if cutfill_d > 0:
            total_cut_volume += cutfill_d
        elif cutfill_d < 0:
            total_fill_volume += abs(cutfill_d)
            
    average_score = (total_cut_volume + total_fill_volume) / 2
    
    return average_score, total_cut_volume, total_fill_volume

# 구조물 판단 및 비용 계산
def cal_profile_structure(new_cutfill):
    new_structure = []

    for station, cutfill_d in new_cutfill:
        if cutfill_d >= 15:
            new_structure.append([station, "터널"])
        elif cutfill_d <= -15:
            new_structure.append([station, "교량"])
        else:
            new_structure.append([station, "토공"])

    return new_structure

# 공사비용 계산 함수
def cal_profile_money(new_structure):
    length_bridge = 0
    length_tunnel = 0
    length_excavation = 0

    for station, structure_type in new_structure:
        if "교량" in structure_type:
            length_bridge += chain
        if "터널" in structure_type:
            length_tunnel += chain
        if "토공" in structure_type:
            length_excavation += chain

    cost_bridge = (length_bridge / 1000) * 24574
    cost_tunnel = (length_tunnel / 1000) * 15784
    cost_excavation = (length_excavation / 1000) * 8620

    new_total_cost = cost_bridge + cost_tunnel + cost_excavation

    return new_total_cost

# 구조물 범위 찾기
def find_structure_ranges(structure_list):
    ranges = {}
    bridge_count = 1
    tunnel_count = 1

    current_structure = None
    start_station = None

    for i, (station, structure) in enumerate(structure_list):
        if structure == '교량':
            if current_structure != '교량':
                if current_structure is not None:
                    if current_structure == '터널':
                        key = f'T{tunnel_count}'
                        ranges[key] = (start_station, structure_list[i - 1][0])
                        tunnel_count += 1
                start_station = station
                current_structure = '교량'
        elif structure == '터널':
            if current_structure != '터널':
                if current_structure is not None:
                    if current_structure == '교량':
                        key = f'B{bridge_count}'
                        ranges[key] = (start_station, structure_list[i - 1][0])
                        bridge_count += 1
                start_station = station
                current_structure = '터널'
        else:
            if current_structure == '교량':
                key = f'B{bridge_count}'
                ranges[key] = (start_station, structure_list[i - 1][0])
                bridge_count += 1
            elif current_structure == '터널':
                key = f'T{tunnel_count}'
                ranges[key] = (start_station, structure_list[i - 1][0])
                tunnel_count += 1
            current_structure = None

    if current_structure == '교량':
        key = f'B{bridge_count}'
        ranges[key] = (start_station, structure_list[-1][0])
    elif current_structure == '터널':
        key = f'T{tunnel_count}'
        ranges[key] = (start_station, structure_list[-1][0])

    return ranges

# 구조물 갯수 세기
def count_structures(structure_ranges):
    bridge_count = 0
    tunnel_count = 0

    for key in structure_ranges:
        if key.startswith('B'):
            bridge_count += 1
        elif key.startswith('T'):
            tunnel_count += 1

    return bridge_count, tunnel_count

# 계획고 plot
# Plotting and combo box integration
def plot_profile(profile, gl, label, color):
    station_elv = generate_station_elv(profile)
    stations = [station for station, elevation in profile]
    elevations = [elevation for station, elevation in profile]
    
    ax.plot(stations, elevations, label=label, color=color)

    #기울기 표시
    # 각 계획선의 중심에 기울기 표시
    i= 0
    for i in range(len(stations) - 1):
        x1, y1 = stations[i], elevations[i]
        x2, y2 = stations[i + 1], elevations[i + 1]
        
        
        gradient = (y2 - y1) / (x2 - x1) * 1000 # 기울기 계산
        if gradient !=0:
            gradient_text = f'{gradient:.2f}'.rstrip('0').rstrip('.')
        else:
            gradient_text = 'L'  # 분모가 0인 경우 'L'로 설정
        
        midpoint = (x1 + x2) / 2, 50# 두 점 사이의 중심

        ax.text(midpoint[0], midpoint[1], gradient_text, fontsize=15, color=color, ha='center', va='center')

    cutfill = calculate_cutfill(station_elv, gl)
    # Calculate and print construction cost
    struct = cal_profile_structure(cutfill)
    # Determine and print structures
    struct_station = find_structure_ranges(struct)
    
    #구조물 시종점 표시
    # 교량 시종점에 수직선 추가
    for key, (start_station, end_station) in struct_station.items():
        if start_station == end_station:
            end_station = start_station + chain
        for station, elevation in station_elv:
            if key.startswith('B') and start_station == station:
                ax.plot([start_station, start_station], [elevation, elevation + 100], color=color, linestyle='-')
                ax.text(start_station, elevation + 100, key, color=color)
            elif key.startswith('B') and end_station == station:
                ax.plot([end_station, end_station], [elevation, elevation + 100], color=color, linestyle='-')

            elif key.startswith('T') and start_station == station:
                ax.plot([start_station, start_station], [elevation, elevation + 100], color=color, linestyle='-')
                ax.text(start_station, elevation + 100, key, color=color)
            elif key.startswith('T') and end_station == station:
                ax.plot([end_station, end_station], [elevation, elevation + 100], color=color, linestyle='-')
            else:
                continue
    #VIP종선표시
    for station, elevation in profile:
        diff_elev = elevation - 0
        x1, y1 = station, elevation
        x2, y2 = station, elevation - diff_elev
        ax.plot([station, station], [elevation, elevation - diff_elev], color=color, linestyle='-')
        ax.text(station, elevation - diff_elev + 40 , f'{elevation:.2f}',fontsize=15, color=color, ha='center',
                va='center',rotation='vertical',bbox=dict(facecolor='white', alpha=1, edgecolor='red'))
    
     # 첫 번째 측점과 마지막 측점을 연결하는 수평선 플롯
    first_point_elevation = 0
    last_point_elevation = 0
    ax.plot([profile[0][0], profile[-1][0]], [first_point_elevation, first_point_elevation], color='k', linestyle='--')


def plot_ground_profile(gl):
    stations = [station for station, elevation in gl]
    elevations = [elevation for station, elevation in gl]
    
    ax.plot(stations, elevations, label='지반고', color='black')
    
# 엑셀파일 생성 함수
def create_excel(alternative,name):

    # 엑셀 파일 생성
    wb = openpyxl.Workbook()

    # '종단' 시트 생성
    ws1 = wb.active
    ws1.title = "종단"

    # 데이터 쓰기
    # 헤더 추가
    ws1.append(["측점", "계획고"])
    for station, elevation in alternative['종단']:
        ws1.append([station, elevation])

    # '구조물' 시트 생성
    ws2 = wb.create_sheet(title="교량")
    ws3 = wb.create_sheet(title="터널")
    # 데이터 쓰기
    # 헤더 추가
    ws2.append(["명칭", "시점", "종점","길이","시작구문","종점구문"])
    ws3.append(["명칭", "시점", "종점","길이","시작구문","종점구문"])
    
    for struct_type, (start_station, end_station) in alternative['구조물위치'].items():
        length = end_station - start_station
        if length ==0:
            length = chain
            end_station = start_station + length
        if struct_type.startswith('B'):
            bridge_start_bve = str(start_station) + ",.wall 0;0;0;,.dikeend 0;"
            bridge_end_bve = str(end_station) + ",.wallend 0;,.dike 0;0;32;"
            ws2.append([struct_type, start_station, end_station,length,bridge_start_bve,bridge_end_bve])
        else:
            tunnel_start_bve = str(start_station) + ",.wall 0;-1;51;,.dikeend 0;"
            tunnel_end_bve = str(end_station) + ",.wallend 0;,.dike 0;0;32;"
            ws3.append([struct_type, start_station, end_station,length,tunnel_start_bve,tunnel_end_bve])
    
    # '절성고' 시트 생성
    ws4 = wb.create_sheet(title="절성고")

    # 데이터 쓰기
    # 헤더 추가
    ws4.append(["측점", "절성토"])
    
    for station,diff in alternative['절성고']:
        ws4.append([station, diff*-1])
    
    # 엑셀 파일 저장
    dierctory = 'c:\\temp\\'
    filename = name + '.xlsx'
    
    wb.save(dierctory + filename)

def print_profile(profile,name):
    print("\n-----Selected Profile Information:-----\n")
    print(f'{name}')
    # Calculate and print score
    station_elv = generate_station_elv(profile)
    cutfill = calculate_cutfill(station_elv, gl)
    score, total_cut_volume, total_fill_volume = calculate_cutfill_volume_and_score(cutfill)
    print(f"Score: {score}")
    
    # Calculate and print construction cost
    struct = cal_profile_structure(cutfill)
    
    construction_cost = cal_profile_money(struct)  # You need to define this function
    print(f"공사비 : {construction_cost:.2f}백만원")
    
    # Calculate and print steepest gradient
    steepest_gradient, steepest_gradient_location = find_steepest_gradient(profile)
    print(f'최급기울기 = {steepest_gradient:.2f} 퍼밀')
    print(f'기울기개소 = {steepest_gradient_location} 개소')

    # Determine and print structures
    struct_station = find_structure_ranges(struct)
    count_struct = count_structures(struct_station)
    
    print(f'구조물 = (교량: {count_struct[0]}개소, 터널: {count_struct[1]}개소)')

def toggle_aspect_ratio():
    if aspect_var.get() == 1:
        ax.set_aspect((1000/400), adjustable='datalim')
    else:
        ax.set_aspect('auto', adjustable='box')
    canvas.draw()

def profile_spec(profile):
    alternative = {}
    alternative['종단'] = profile
    alternative['계획고'] = generate_station_elv(alternative['종단']) 
    alternative['절성고'] = calculate_cutfill(alternative['계획고'], gl)
    alternative['구조물'] = cal_profile_structure(alternative['절성고'])
    alternative['구조물위치'] = find_structure_ranges(alternative['구조물'])
    alternative['구조물갯수'] = count_structures(alternative['구조물위치'])
    alternative['공사비'] = cal_profile_money(alternative['구조물'])
    alternative['점수'] = calculate_cutfill_volume_and_score(alternative['절성고'])
    
    return alternative

gl = load_gl()
profile = load_fl()

if gl and profile:
    station_elv = generate_station_elv(profile)
    new_cutfill = calculate_cutfill(station_elv, gl)
    new_structure = structure_judgement(new_cutfill)
    average_score, total_cut_volume, total_fill_volume = calculate_cutfill_volume_and_score(new_cutfill)

# GUI 구성
root = tk.Tk()
root.title("종단면도")

# Create a variable to store the checkbox state
aspect_var = tk.IntVar(value=0)

# Create the checkbox
aspect_checkbox = tk.Checkbutton(root, text="비율 동일하게", variable=aspect_var, command=toggle_aspect_ratio)
aspect_checkbox.pack()

# Initialize the plot
fig = plt.figure(figsize=(6, 6))
ax = fig.add_subplot()
canvas = FigureCanvasTkAgg(fig, master=root)
canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

# Create a frame for the toolbar
toolbar_frame = tk.Frame(root)
toolbar_frame.pack(side=tk.BOTTOM, fill=tk.X)

# Add the navigation toolbar
toolbar = NavigationToolbar2Tk(canvas, toolbar_frame)
toolbar.update()

# Run the GUI
root.geometry("900x800")

plot_profile(profile, gl, '계획안','red')
plot_ground_profile(gl)
print_profile(profile,'계획안')
alternative = profile_spec(profile)
create_excel(alternative,'최적대안')

print(len(gl))
print(len(station_elv))
#txt파일로 저장
# gl 저장
with open('c:/temp/gl.txt', 'w', newline='') as f:
    writer = csv.writer(f, delimiter='\t')
    writer.writerows(gl)

# station_elv 저장
with open('c:/temp/station_elv.txt', 'w', newline='') as f:
    writer = csv.writer(f, delimiter='\t')
    writer.writerows(station_elv)
print('실행완료')
root.mainloop()
